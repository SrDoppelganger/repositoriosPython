import pandas as pd
import openpyxl

municipios_aceitos: list[str] = []
escolas_aceitas: list[str] = []
topicos_aceitos: list[str] = []
subtopicos_aceitos: list[str] = []
links_aceitos: list[str] = []

def processar_planilha(sheets):
    for sheet in sheets:
        print(f"Trabalhando na aba {sheet}...")
        df = pd.read_excel(input_file, sheet_name=sheet,header=1)
        processar_municipio(df,sheet)

    gerar_planilha()

def processar_municipio(df,sheet):
    #arrays de colunas
    link_cols: list[int] = [7,10,13,16,19,22]
    check_cols: list[int] = [9,12,15,18,21,24]

    for i in range(6):
        escola_values = df[df.columns[1]]
        topico_values = df[df.columns[5]]
        subtopico_values = df[df.columns[6]]

        link_values = df[df.columns[link_cols[i]]]
        check_values = df[df.columns[check_cols[i]]]

        for index,check in enumerate(check_values):
            if check == True or check == "TRUE":
                municipios_aceitos.append(sheet)
                escolas_aceitas.append(escola_values[index])
                topicos_aceitos.append(topico_values[index])
                subtopicos_aceitos.append(subtopico_values[index])
                links_aceitos.append(link_values[index])

def gerar_planilha():
    # Criação da planilha
    wb = openpyxl.Workbook()
    wb.create_sheet('links_aceitos')
    wb.remove(wb['Sheet'])
    wb['links_aceitos'].append(['Município','Unidade_Administrativa','Tópico','Subtópico','Links'])
    ws = wb.active

    for row in zip(municipios_aceitos,escolas_aceitas, topicos_aceitos, subtopicos_aceitos, links_aceitos):
        ws.append(row)

    wb.save("Links aceitos.xlsx")

# Função principal
if __name__ == "__main__":

    input_file = "selecao.xlsx"
    wb = openpyxl.open(input_file)
    sheets = wb.sheetnames
    processar_planilha(sheets)
