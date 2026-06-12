#pega TODAS as fotos da base e coloca no formato da planilha do painel

import pandas as pd
import openpyxl

auditores_aceitos: list[str] = []
municipios_aceitos: list[str] = []
escolas_aceitas: list[str] = []
inep_aceitos:list [str] = []
questoes_aceitas:list [str] = []
cod_questoes_aceitas:list [str] = []
topicos_aceitos: list[str] = []
links_aceitos: list[str] = []

def processar_planilha(sheets):
    for sheet in sheets:
        print(f"Trabalhando na aba {sheet}...")
        df = pd.read_excel(input_file, sheet_name=sheet,header=1)
        processar_municipio(df,sheet)

    gerar_planilha()

def processar_municipio(df,sheet):
    #arrays de colunas
    link_cols: list[int] = [7,10,13,16,19,22]
    check_cols: list[int] = [9,12,15,18,21,24]

    for i in range(6):
        auditor_values = df[df.columns[0]]
        escola_values = df[df.columns[1]]
        cod_questao_values = df[df.columns[2]]
        questao_values = df[df.columns[3]]
        topico_values = df[df.columns[5]]
        link_values = df[df.columns[link_cols[i]]]
        check_values = df[df.columns[check_cols[i]]]

        for index,check in enumerate(check_values):
            #checar NaN
            if not pd.isna(check) and link_values[index] != "-":
                auditores_aceitos.append(auditor_values[index])
                municipios_aceitos.append(sheet)
                escolas_aceitas.append(remover_inep(escola_values[index]))
                inep_aceitos.append(pegar_inep(escola_values[index]))
                cod_questoes_aceitas.append(cod_questao_values[index])
                questoes_aceitas.append(questao_values[index])
                topicos_aceitos.append(topico_values[index])
                links_aceitos.append(link_values[index])

def pegar_inep(cell):
    #evita erros com linhas vazias T .T
    if type(cell) is str:
        start = cell.find('(') + 1
        end = cell.find(')', start)

        if start > 0 and end > -1:
            text = cell[start:end]
            return text

def remover_inep(cell):
    #evita erros com linhas vazias T .T
    if type(cell) is str:
        text = cell.split("(")[0]
        return text

def gerar_planilha():
    # Criação da planilha
    wb = openpyxl.Workbook()
    wb.create_sheet('links_aceitos')
    wb.remove(wb['Sheet'])
    wb['links_aceitos'].append(['Auditor','MUNICIPIO -  ','UNIDADE_ADMINISTRATINA -  ','SIGLA_UNIDADE_ADM -  ','Questão-Código','Questão','Tópico','Links'])
    ws = wb.active

    for row in zip(auditores_aceitos,municipios_aceitos,escolas_aceitas, inep_aceitos, cod_questoes_aceitas,questoes_aceitas,topicos_aceitos, links_aceitos):
        ws.append(row)

    wb.save("extrato_fotos.xlsx")

# Função principal
if __name__ == "__main__":

    input_file = "selecao.xlsx"
    wb = openpyxl.open(input_file)
    sheets = wb.sheetnames
    processar_planilha(sheets)