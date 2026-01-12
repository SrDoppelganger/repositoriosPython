#código de calculo de TAGS refatorado
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import argparse
import re
import unicodedata
from pathlib import Path
from typing import Dict, List, Set, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from functools import reduce
import operator

# ========================= 1. CONFIGURAÇÃO CENTRALIZADA =========================
# Mapeamento de nomes amigáveis para colunas do Excel.
# Este é o ÚNICO lugar para configurar colunas.
COLS_AS = {
    "MUNICIPIO": "F", "INEP": "G", "ESFERA": "K", "ACESSO": "Z",
    "ETAPAS": "AJ","ACEI":"BR", "ESGOTAMENTO": "EE", "DEPENDENCIAS": "ER", "DEP_INF": "ET",
    "IRR_ENTRADA": "AE", "GEST_ALIM": "BL", "TERC_ALIM": "BN","EST_CONT":"BW","EST_CONF":"CF","ANVISA": "CK",
    "AVCB": "CP", "DEDET": "CU", "ABS_AGUA": "CZ", "RES_AGUA": "DB","CAIXA_LIMP":"DF", "CIST_LIMP":"DJ", "ORIGEM_AGUA":"DN",
    "AGUA_CERTIF":"DS","ABAST_DESC":"DZ","AGUA_POT":"DX","LIXO": "EK", "ENERGIA": "EP", "PATIO": "EV", "PARQ_INF": "FA","BIB_SL":"FK","QUAD_FUNC":"FF",
    "IRR_BIB": "FV", "IRR_SL": "GA", "IRR_BIBSL": "GF", "AG_SAN_EI": "GN",
    "IRR_SAN_EI": "GV","BG_AGUA":"HD", "IRR_BAN": "HK","SAN_PCD":"HP","BAN_PCD_AG": "HV","ITEM_PCD":"IA","IRR_BAN_PCD": "ID",
    "MULTS": "II", "SALA_IRR": "IU","LAV_AGUA":"JT","LAV_ITEM":"JV", "LACT_AGUA":"KC","ITEM_LACT": "KE", "IRR_LACT": "KJ","BERC_ROUPA":"JJ",
    "LOCAL_LACT": "JX", "LOCAL_FRAL": "KO", "ITEM_FRAL": "KT", "IRR_FRAL": "KV",
    "EQP_COZ": "LF", "COZ_OUT": "LR", "LOC_ARM": "LW", "LIM_BERC": "JD",
    "IRR_BER": "JL","COZ_AGUA":"LD","COZ_BALANCA":"LH", "IRR_COZ": "LM", "ARM_IRR": "MB", "ALM_IRR": "MG",
    "ALM_UP": "ML", "ALM_CONG": "MV", "IRR_CONG": "NA", "REC_ADQ":"NM","QTD_ADQ":"NO","FALTA_MERENDA":"NQ","OFERTA_FRUTA":"NS","OFERTA_LEGUME":"NU","CARD": "NW",
    "CARD_ESP": "OB", "REF_SERV": "OG", "REF_CONF": "OM", "ING_REF": "OR",
    "MOB_REF": "OW", "IRR_REF": "PB", "CONS_UP": "PG", "VEND_UP": "PL", "DIR_UP": "PR","RISCO_GRAVE":"PX"
}

# Formato: 'chave_interna': {'col': 'NOME_DA_COLUNA', 'text': 'TEXTO_A_PROCURAR', 'tag': '<<TAG_BASE>>'}
CALCULATION_SPECS = {
    # Escolas Visitadas
    # must_contain permite que a string esteja no meio de uma lista
    'todas': {'col': 'INEP', 'op': 'notna', 'tag': '<<NUMERO_ESCOLAS_VISITADAS>>'},
    'infantil_q': {'col': 'ETAPAS', 'op':'composite','conditions':{'must_contain':['Educação Infantil']}, 'tag': '<<ESCOLAS_INFANTIL_VISITADAS>>'},
    'infantil_creche': {'col': 'ETAPAS', 'op':'composite','conditions':{'must_contain':['Educação Infantil - Creche']}, 'tag': '<<ESCOLAS_INFANTIL_VISITADAS_CRECHE>>'},
    'infantil_pre': {'col': 'ETAPAS','op':'composite','conditions':{'must_contain':['Educação Infantil - Pré-escola']}, 'tag': '<<ESCOLAS_INFANTIL_VISITADAS_PREESCOLA>>'},
    'fund_q': {'col': 'ETAPAS', 'op':'composite','conditions':{'must_contain':['Fundamental']}, 'tag': '<<ESCOLAS_FUND_VISITADAS>>'},
    'medio_q': {'col': 'ETAPAS', 'op':'composite','conditions':{'must_contain':['Ensino Médio']}, 'tag': '<<ESCOLAS_MED_VISITADAS>>'},
    'eja_q': {'col': 'ETAPAS', 'op':'composite','conditions':{'must_contain':['EJA']}, 'tag': '<<ESCOLAS_EJA_VISITADAS>>'},
    'inf_fund_exclusivo': {'col': 'ETAPAS', 'op':'composite','conditions':{'must_contain':['Educação Infantil', 'Fundamental'], 'must_not_contain':['Ensino Médio']},'tag':'<<ESCOLAS_INF_FUND_VISITADAS>>'},
    
    # Irregularidades Entrada
    'sem_rampas': {'col': 'IRR_ENTRADA','op':'composite','conditions':{'must_contain':['Não há rampa de acesso, mesmo ela sendo necessária']}, 'tag': '<<ESCOLAS_SEM_RAMPAS>>'},
    'rampas_irregulares': {'col': 'IRR_ENTRADA', 'op':'composite','conditions':{'must_contain':['Há rampa de acesso, mas ela apresenta alguma irregularidade']}, 'tag': '<<ESCOLAS_RAMPAS_IRREGULARES>>'},
    'sem_vao_entrada': {'col': 'IRR_ENTRADA', 'op':'composite','conditions':{'must_contain':['Não há porta de entrada com largura de vão livre igual ou superior a 80cm']}, 'tag': '<<ESCOLAS_VAO_ENTRADA_IRREGULAR>>'},

    'controle_estoque':{'col':'EST_CONT','text':'Sim','tag':'<<ESTOQUE_CONTROLE_SIM>>'},
    'controle_estoque_nao':{'col':'EST_CONT','text':'Não','tag':'<<ESTOQUE_CONTROLE_NAO>>'},

    'estoque_confirmacao_todos':{'col':'EST_CONF','text':'Sim','tag':'<<ESTOQUE_CONF_TODOS>>'},
    'estoque_confirmacao_alguns':{'col':'EST_CONF','text':'Não, mas existe de outros gêneros alimentícios (frutas e legumes, por exemplo)','tag':'<<ESTOQUE_CONF_ALGUNS>>'},
    'estoque_confirmacao_nenhum':{'col':'EST_CONF','text':'Não existe nota de recebimento de gêneros alimentícios na escola','tag':'<<ESTOQUE_CONF_NENHUM>>'},

    # Vigilância Sanitária
    'com_anvisa': {'col': 'ANVISA', 'text': 'Sim, válida', 'tag': '<<LIC_ANVISA_VALIDO>>'},
    'anvisa_fv': {'col': 'ANVISA', 'text': 'fora da validade', 'tag': '<<LIC_ANVISA_FORA_VAL>>'},
    'sem_anvisa': {'col': 'ANVISA', 'text': 'Não', 'tag': '<<LIC_ANVISA_NAO>>'},
    'com_avcb': {'col': 'AVCB', 'text': 'Sim, válida', 'tag': '<<AVCB_VALIDO>>'},
    'avcb_fv': {'col': 'AVCB', 'text': 'fora da validade', 'tag': '<<AVCB_FORA_VAL>>'},
    'sem_avcb': {'col': 'AVCB', 'text': 'Não', 'tag': '<<AVCB_NAO>>'},
    'dedet_dp': {'col': 'DEDET', 'text': 'Sim, emitido há no máximo 6 meses', 'tag': '<<DEDET_ATE6M>>'},
    'dedet_fp': {'col': 'DEDET', 'text': 'Sim, emitido há mais de 6 meses', 'tag': '<<DEDET_MAIS6M>>'},
    'sem_dedet': {'col': 'DEDET', 'text': 'Não', 'tag': '<<DEDET_NAO>>'},
    
    # Abastecimento de Água
    'rede_publica': {'col': 'ABS_AGUA','op':'composite','conditions':{'must_contain':['Rede Pública']},'tag': '<<AGUA_REDE_PUBLICA>>'},
    'poco_artesiano': {'col': 'ABS_AGUA','op':'composite','conditions':{'must_contain':['Poço artesiano']}, 'tag': '<<AGUA_POCO_ARTESIANO>>'},
    'cacimba_etc': {'col': 'ABS_AGUA', 'op':'composite','conditions':{'must_contain':['Cacimba/cisterna/poço']}, 'tag': '<<AGUA_CACIMBA_CISTERNA_POCO>>'},
    'fonte_etc': {'col': 'ABS_AGUA', 'op':'composite','conditions':{'must_contain':['Fonte/rio/igarapé/riacho/córrego']}, 'tag': '<<AGUA_FONTE_RIO_IGARAPE_RIACHO_CORREGO>>'},
    'sem_agua': {'col': 'ABS_AGUA', 'op':'composite','conditions':{'must_contain':['Não há abastecimento de água']}, 'tag': '<<AGUA_NAO_HA>>'},

    # Energia elétrica en->energia
    'en_sim_func':{'col':'ENERGIA', 'text':'Sim, em funcionamento', 'tag':'<<ENERGIA_SIM_FUNC>>'},
    'en_sim_fora':{'col':'ENERGIA', 'text':'Sim, mas fora de funcionamento', 'tag':'<<ENERGIA_FORA_FUNC>>'},
    'en_nao_ha':{'col':'ENERGIA', 'text':'Não', 'tag':'<<ENERGIA_NAO>>'},

    #Deps (que não funcionam com tags dinamicas)
    'dep_patio_coberto':{'col':'DEPENDENCIAS', 'op':'composite', 'conditions':{'must_contain':['Pátio coberto']}, 'tag':'<<DEP_PATIO_COBERTO>>'},
    'dep_patio_descoberto':{'col':'DEPENDENCIAS', 'op':'composite', 'conditions':{'must_contain':['Pátio descoberto']}, 'tag':'<<DEP_PATIO_DESCOBERTO>>'},

    #Deps Infantil
    'dep_inf_san':{'col':'DEP_INF', 'op':'composite','conditions':{'must_contain':['Sanitário exclusivo à educação infantil']}, 'tag':'<<DEP_SANITARIO_ED_INFANTIL>>'},
    'dep_inf_san_existe':{'col':'DEP_INF', 'op':'composite','conditions':{'must_contain':['Sanitário exclusivo à educação infantil']}, 'tag':'<<EI_SAN_EXC_EXISTE>>'}, # NÃO TENHO CERTEZA SE É ASSIM QUE SE CALCULA ESSA TAG
    'dep_inf_berc':{'col':'DEP_INF', 'op':'composite','conditions':{'must_contain':['Berçário']}, 'tag':'<<DEP_BERCARIO>>'},
    'dep_inf_fral':{'col':'DEP_INF', 'op':'composite','conditions':{'must_contain':['Fraldário']}, 'tag':'<<DEP_FRALDARIO>>'},
    'dep_inf_lact':{'col':'DEP_INF', 'op':'composite','conditions':{'must_contain':['Lactário']}, 'tag':'<<DEP_LACTARIO>>'},
    'dep_inf_parq':{'col':'DEP_INF', 'op':'composite','conditions':{'must_contain':['Parque infantil']}, 'tag':'<<DEP_PARQUE_INFANTIL>>'},
    'dep_inf_lav':{'col':'DEP_INF', 'op':'composite','conditions':{'must_contain':['Lavanderia e/ou área de lavagem, secagem e armazenamento']}, 'tag':'<<DEP_LAVANDERIA_AREA_LAVAGEM>>'},
    'dep_inf_parq':{'col':'DEP_INF','text':'Não', 'tag':'<<BERCARIO_LIMITE_CRIANÇAS_IRREG>>'},

    # Salas de Aula mult->multisseriada
    'mult_sim':{'col':'MULTS', 'text':'Sim', 'tag':'<<SALAS_MULTISSERIADAS>>'},
    'sala_irr':{'col': 'SALA_IRR', 'op':'composite','conditions':{'must_not_contain':['Não foram identificados os aspectos irregulares listados acima']},'tag':'<<SALAS_IRREGULARES>>'},
    'sala_irr_ilum':{'col': 'SALA_IRR','op':'composite','conditions':{'must_contain':['Ausência de iluminação natural']},'tag':'<<SALAS_IRREGULARES_ILUMINACAO>>'},
    'sala_irr_ventilacao':{'col': 'SALA_IRR','op':'composite','conditions':{'must_contain':['Ausência de ventilação natural']},'tag':'<<SALAS_IRREGULARES_VENTILACAO>>'},
    'sala_irr_ventiladores':{'col': 'SALA_IRR','op':'composite','conditions':{'must_contain':['Ausência de ventiladores e/ou ar condicionado']},'tag':'<<SALAS_IRREGULARES_VENTILADORES>>'},
    'sala_irr_pisos':{'col': 'SALA_IRR','op':'composite','conditions':{'must_contain':['Pisos sujos, com rachaduras, buracos e/ou falhas no revestimento']},'tag':'<<SALAS_IRREGULARES_PISO>>'},
    'sala_irr_teto':{'col': 'SALA_IRR','op':'composite','conditions':{'must_contain':['Teto/cobertura com rachaduras, buracos, quebra de reboco, mofo e/ou infiltrações']},'tag':'<<SALAS_IRREGULARES_TETO>>'},
    'sala_irr_parede':{'col': 'SALA_IRR','op':'composite','conditions':{'must_contain':['Paredes com rachaduras, buracos, quebra de reboco, mofo e/ou infiltrações']},'tag':'<<SALAS_IRREGULARES_PAREDES>>'},
    'sala_irr_cantos':{'col': 'SALA_IRR','op':'composite','conditions':{'must_contain':['Cantos pontiagudos nos equipamentos']},'tag':'<<SALAS_IRREGULARES_CANTOS>>'},

    # Esgotamento
    'sist_conectado': {'col': 'ESGOTAMENTO', 'text': 'Conexão com rede de esgotamento sanitário', 'tag': '<<ESGOTO_REDE_SANITARIA>>'},
    'fossa_e_outros': {'col': 'ESGOTAMENTO', 'text': 'Fossa, sumidouro ou similar', 'tag': '<<ESGOTO_FOSSA_SUMIDOURO>>'},
    'despejo_inadequado': {'col': 'ESGOTAMENTO', 'text': 'Despejo sem destinação adequada', 'tag': '<<ESGOTO_DESPEJO_INADEQUADO>>'},
    
   
    # Alimentação:
    'gest_alim_centralizada': {'col': 'GEST_ALIM', 'op':'composite','conditions':{'must_contain':['Centralizada'], 'must_not_contain':['Descentralizada ou Escolarizada','Semi Descentralizada ou Parcialmente Escolarizada','Terceirizada']}, 'tag': '<<GESTAO_ALIM_CENTRALIZADA>>'},
    'gest_alim_descentralizada': {'col': 'GEST_ALIM', 'op':'composite','conditions':{'must_contain':['Descentralizada ou Escolarizada'], 'must_not_contain':['Centralizada']}, 'tag':'<<GESTAO_ALIM_DESCENTRALIZADA>>'},
    'gest_alim_semi_descentralizada': {'col': 'GEST_ALIM', 'op':'composite','conditions':{'must_contain':['Semi Descentralizada ou Parcialmente Escolarizada'], 'must_not_contain':['Centralizada','Descentralizada ou Escolarizada']}, 'tag':'<<GESTAO_ALIM_SEMIDESCENTRALIZADA>>'},
    'gest_alim_terceirizada': {'col': 'GEST_ALIM', 'op':'composite','conditions':{'must_contain':['Terceirizada'], 'must_not_contain':['Centralizada','Descentralizada ou Escolarizada','Semi Descentralizada ou Parcialmente Escolarizada']}, 'tag':'<<GESTAO_ALIM_TERCEIRIZADA>>'},

    'terc_alim_eqp_escola': {'col': 'TERC_ALIM', 'op':'composite','conditions':{'must_contain':['Com preparo na escola, usando equipamentos da escola']}, 'tag': '<<GESTAO_ALIM_TERCEIRIZADA_PREP_ESCOLA_EQP_ESCOLA>>'},
    'terc_alim_eqp_empresa': {'col': 'TERC_ALIM', 'op':'composite','conditions':{'must_contain':['Com preparo na escola, usando equipamentos da empresa']}, 'tag': '<<GESTAO_ALIM_TERCEIRIZADA_PREP_ESCOLA_EQP_EMPRESA>>'},
    'terc_alim_hotbox': {'col': 'TERC_ALIM', 'op':'composite','conditions':{'must_contain':['Com entrega em hot box']}, 'tag': '<<GESTAO_ALIM_TERCEIRIZADA_ENTREGA_HOTBOX>>'},

    'tem_card_esp': {'col': 'CARD_ESP','op':'composite','conditions':{'must_contain':['Sim, havendo cardápio especial para esses alunos'], 'must_not_contain':['Sim, mas não havendo cardápio especial para esses alunos','Não']}, 'tag': '<<CARDAPIO_NEEDS_ESPECIAL_COM_CARDAPIO>>'},
    'sem_card_esp': {'col': 'CARD_ESP','op':'composite','conditions':{'must_contain':['Sim, mas não havendo cardápio especial para esses alunos'], 'must_not_contain':['Sim, havendo cardápio especial para esses alunos','Não']}, 'tag': '<<CARDAPIO_NEEDS_ESPECIAL_SEM_CARDAPIO>>'},
    'card_esp_nao': {'col': 'CARD_ESP','op':'composite','conditions':{'must_contain':['Não'], 'must_not_contain':['Sim, havendo cardápio especial para esses alunos','Sim, mas não havendo cardápio especial para esses alunos']}, 'tag': '<<CARDAPIO_NEEDS_NAO>>'},

    'ref_serv': {'col': 'REF_SERV', 'text': 'Sim, havia refeição sendo SERVIDA', 'tag': '<<CARDAPIO_VISITA_REFEICAO_SERVIDA>>'},
    'ref_prep': {'col': 'REF_SERV', 'text': 'Sim, havia refeição sendo PREPARADA', 'tag': '<<CARDAPIO_VISITA_REFEICAO_PREPARADA>>'},
    'ref_nao': {'col': 'REF_SERV', 'text': 'Não', 'tag': '<<CARDAPIO_VISITA_REFEICAO_NAO>>'},

    'estoque_ing_todos': {'col': 'ING_REF', 'text': 'Sim, todos os ingredientes necessários para todas as refeições daquele dia', 'tag': '<<CARDAPIO_ESTOQUE_ING_TODOS>>'},
    'estoque_ing_alguns': {'col': 'ING_REF', 'text': 'Sim, os ingredientes necessários para algumas das refeições daquele dia', 'tag': '<<CARDAPIO_ESTOQUE_ING_ALGUNS>>'},
    'estoque_ing_nenhum': {'col': 'ING_REF', 'text': 'Não, nenhum dos ingredientes necessários para as refeições daquele dia', 'tag': '<<CARDAPIO_ESTOQUE_ING_NENHUM>>'},

    'coz_agua_reg': {'col': 'COZ_AGUA', 'text': 'Sim', 'tag': '<<COZ_AGUA_REGULAR_SIM>>'},
    'coz_agua_irr': {'col': 'COZ_AGUA', 'text': 'Não', 'tag': '<<COZ_AGUA_REGULAR_NAO>>'},
    'coz_balanca': {'col': 'COZ_BALANCA', 'text': 'Sim', 'tag': '<<COZ_BALANCA_SIM>>'},
    'coz_balanca_irr': {'col': 'COZ_BALANCA', 'text': 'Não', 'tag': '<<COZ_BALANCA_NAO>>'},
    #Fraldário
    'local_dentro': {'col': 'LOCAL_FRAL', 'text': 'Implantado dentro do berçário', 'tag': '<<FRALDARIO_LOCAL_DENTRO_BERCARIO>>'},
    'local_separado': {'col': 'LOCAL_FRAL', 'text': 'Em ambiente separado', 'tag': '<<FRALDARIO_LOCAL_SEPARADO>>'},

    #Berçário
    'berc_roupa_cama': {'col': 'BERC_ROUPA', 'text': 'Não', 'tag': '<<BERCARIO_ROUPA_CAMA_IRREG>>'},

    #Lavanderia
    'lav_tanque': {'col': 'LAV_ITEM', 'op':'composite','conditions':{'must_contain':['Tanque']}, 'tag': '<<LAVANDERIA_TANQUE>>'},
    'lav_maquina': {'col': 'LAV_ITEM', 'op':'composite','conditions':{'must_contain':['Máquina de lavar e/ou secar']}, 'tag': '<<LAVANDERIA_MAQUINA>>'},
    'lav_varal': {'col': 'LAV_ITEM', 'op':'composite','conditions':{'must_contain':['Varal']}, 'tag': '<<LAVANDERIA_VARAL>>'},
    'lav_material': {'col': 'LAV_ITEM', 'op':'composite','conditions':{'must_contain':['Material de limpeza']}, 'tag': '<<LAVANDERIA_MATERIAL>>'},
    'lav_nenhum_item': {'col': 'LAV_ITEM', 'op':'composite','conditions':{'must_contain':['Nenhum dos itens elencados']}, 'tag': '<<LAVANDERIA_NENHUM>>'},
    'lav_agua_sim': {'col': 'LAV_AGUA', 'text':'Sim', 'tag': '<<LAVANDERIA_AGUA_SIM>>'},
    'lav_agua_nao': {'col': 'LAV_AGUA', 'text':'Não', 'tag': '<<LAVANDERIA_AGUA_NAO>>'},

    #Lactário
    'lact_agua_sim': {'col': 'LACT_AGUA', 'text':'Sim', 'tag': '<<LACTARIO_AGUA_SIM>>'},
    'lact_agua_nao': {'col': 'LACT_AGUA', 'text':'Não', 'tag': '<<LACTARIO_AGUA_NAO>>'},


    #Infra pátio
    'patio_cob_exclusivo': {'col': 'PATIO', 'text': 'Sim, área COBERTA com horário de utilização EXCLUSIVO para o ensino infantil', 'tag': '<<EDUC_INF_PATIO_COBERTO_EXCLUSIVO>>'},
    'patio_descob_exclusivo': {'col': 'PATIO', 'text': 'Sim, área DESCOBERTA com horário de utilização EXCLUSIVO para o ensino infantil', 'tag': '<<EDUC_INF_PATIO_DESCOBERTO_EXCLUSIVO>>'},
    'patio_cob_compartilhado': {'col': 'PATIO', 'text': 'Sim, área COBERTA COMPARTILHADA no mesmo horário com outras etapas de ensino', 'tag': '<<EDUC_INF_PATIO_COBERTO_COMPART>>'},
    'patio_descob_compartilhado': {'col': 'PATIO', 'text': 'Sim, área DESCOBERTA COMPARTILHADA no mesmo horário com outras etapas de ensino', 'tag': '<<EDUC_INF_PATIO_DESCOBERTO_COMPART>>'},
    'nao_tem_patio':{'col':'PATIO', 'text':'Não', 'tag':'<<EDUC_INF_PATIO_NAO>>'},

    #Parque Infantil
    'parque_tem_cond':{'col':'PARQ_INF', 'text':'Sim, em condições de uso', 'tag':'<<EDUC_INF_PARQUINHO_COND_USO>>'},
    'parque_sem_cond':{'col':'PARQ_INF', 'text':'Sim, mas sem condições de uso', 'tag':'<<EDUC_INF_PARQUINHO_SEM_COND>>'},
    'parque_nao':{'col':'PARQ_INF', 'text':'Não', 'tag':'<<EDUC_INF_PARQUINHO_NAO>>'},

    #Local Armazenamento dos gen. alimenticios
    'local_desp':{'col':'LOC_ARM', 'text':'Em local especifico de despensa', 'tag':'<<ARMZ_LOCAL_DESPENSA>>'},
    'local_arm_dentro':{'col':'LOC_ARM', 'text':'Em armário, dentro na cozinha', 'tag':'<<ARMZ_LOCAL_ARMARIO_DENTRO_COZINHA>>'},
    'local_arm_fora':{'col':'LOC_ARM', 'text':'Em armário, fora do ambiente da cozinha', 'tag':'<<ARMZ_LOCAL_ARMARIO_FORA_COZINHA>>'},
    'local_nao_ha':{'col':'LOC_ARM', 'text':'Não há local de armazenamento de gêneros alimentícios na escola', 'tag':'<<ARMZ_LOCAL_NAO_HA>>'},

    #Cardápio Assinado
    'card_ass_rt':{'col':'CARD', 'text':'Sim, assinado por Nutricionista RT e fixado em local visível', 'tag':'<<CARDAPIO_ASS_RT_E_FIXADO>>'},
    'card_nao_ass':{'col':'CARD', 'text':'Sim, NÃO assinado por nutricionista, mas fixado em local visível', 'tag':'<<CARDAPIO_NAO_ASSINADO_MAS_FIXADO>>'},
    'card_ass_nao_fix':{'col':'CARD', 'text':'Sim, assinado por nutricionista, mas NÃO fixado em local visível', 'tag':'<<CARDAPIO_ASSINADO_NAO_FIXADO>>'},
    'card_nao_ass_nao_fix':{'col':'CARD', 'text':'Sim, NÃO assinado por nutricionista e NÃO fixado em local visível', 'tag':'<<CARDAPIO_NAO_ASS_NAO_FIXADO>>'},
    'card_nao_existe':{'col':'CARD', 'text':'Não', 'tag':'<<CARDAPIO_NAO_EXISTE>>'},

    #TODO: Irregularidades local de armazenamento dos alimentos
    'armz_irr_prat_enf':{'col':'ARM_IRR', 'op':'composite','conditions':{'must_contain':['Prateleira']}, 'tag':'<<ARMZ_LOCAL_IRREG_PRATELEIRA_ENFERRUJADA>>'},
    'armz_irr_sup':{'col':'ARM_IRR', 'op':'composite','conditions':{'must_contain':['Alimentos empilhados sobre superfícies não apropriadas']}, 'tag':'<<ARMZ_LOCAL_IRREG_SUPERFICIE_INAPROPRIADA>>'},
    'armz_irr_chao':{'col':'ARM_IRR', 'op':'composite','conditions':{'must_contain':['Armazenamento de alimentos diretamente no chão']}, 'tag':'<<ARMZ_LOCAL_IRREG_DIRETO_NO_CHAO>>'},
    'armz_irr_mat_limpeza':{'col':'ARM_IRR', 'op':'composite','conditions':{'must_contain':['Armazenamento de materiais de limpeza e/ou outros materiais junto de alimentos']}, 'tag':'<<ARMZ_LOCAL_IRREG_MATERIAIS_LIMPEZA_JUNTO>>'},
    'armz_irr_ar':{'col':'ARM_IRR', 'op':'composite','conditions':{'must_contain':['Ambiente não possui entrada de ar']}, 'tag':'<<ARMZ_LOCAL_IRREG_SEM_ENTRADA_AR>>'},
    'armz_irr_animais':{'col':'ARM_IRR', 'op':'composite','conditions':{'must_contain':['Presença de animais']}, 'tag':'<<ARMZ_LOCAL_IRREG_PRESENCA_ANIMAIS>>'},
    
    'armz_alm_fora_val':{'col':'ALM_IRR', 'op':'composite','conditions':{'must_contain':['Alimentos fora do prazo de validade']}, 'tag':'<<ARMZ_ALIM_FORA_VALIDADE>>'},
    'armz_alm_sem_etq':{'col':'ALM_IRR', 'op':'composite','conditions':{'must_contain':['Ausência de etiquetas de identificação de validade']}, 'tag':'<<ARMZ_ALIM_SEM_ETIQUETA_VAL>>'},
    'armz_alm_mofados':{'col':'ALM_IRR', 'op':'composite','conditions':{'must_contain':['Alimentos visivelmente mofados ou podres']}, 'tag':'<<ARMZ_ALIM_MOFADOS_PODRES>>'},
    'armz_alm_emb_danif':{'col':'ALM_IRR', 'op':'composite','conditions':{'must_contain':['Embalagens abertas, rasgadas ou danificadas expondo os alimentos à contaminação']}, 'tag':'<<ARMZ_ALIM_EMBALAGENS_DANIFICADAS>>'},

    'doc_aceitabilidade':{'col':'ACEI', 'text':'Sim', 'tag':'<<ACEITABILIDADE_SIM>>'},
    'doc_aceitabilidade_nao':{'col':'ACEI', 'text':'Não', 'tag':'<<ACEITABILIDADE_NAO>>'},

    'quadra_func':{'col':'QUAD_FUNC', 'text':'Sim, em estado adequado de conservação e funcionamento', 'tag':'<<QUADRA_FUNC_ADEQ>>'},
    'quadra_func_irr':{'col':'QUAD_FUNC', 'text':'Sim, mas apresenta irregularidades', 'tag':'<<QUADRA_FUNC_IRREG>>'},
    'quadra_nao_func':{'col':'QUAD_FUNC', 'text':'Não', 'tag':'<<QUADRA_NAO_FUNC>>'},

    # BG -> Banheiro Geral
    'bg_agua_reg':{'col':'BG_AGUA', 'text':'Sim', 'tag':'<<BANH_GERAL_AGUA_REGULAR>>'},

    'san_pcd_exc':{'col':'SAN_PCD', 'text':'Sim, em espaço físico EXCLUSIVO', 'tag':'<<PCD_SAN_EXCLUSIVO>>'},
    'san_pcd_mesmo':{'col':'SAN_PCD', 'text':'Sim, no MESMO ESPAÇO físico destinado ao', 'tag':'<<PCD_SAN_MESMO_ESPACO>>'},
    'san_pcd_inex':{'col':'SAN_PCD', 'text':'Não', 'tag':'<<PCD_SAN_INEXISTENTE>>'},

    'biblioteca_exc':{'col':'BIB_SL','op':'composite','conditions':{'must_contain':['Espaço físico exclusivo para biblioteca']} , 'tag':'<<BIBLIOTECA_EXCLUSIVA>>'},
    'sala_leitura_exc':{'col':'BIB_SL','op':'composite','conditions':{'must_contain':['Espaço físico exclusivo para sala de leitura']} , 'tag':'<<SALA_LEITURA_EXCLUSIVA>>'},
    'bib_sl_juntos':{'col':'BIB_SL','op':'composite','conditions':{'must_contain':['Mesmo espaço físico para biblioteca e sala de leitura']} , 'tag':'<<BIBLI_SALA_MESMO_ESPACO>>'},

    'limp_caixa_6m':{'col':'CAIXA_LIMP', 'text':'Sim, nos últimos 6 meses', 'tag':'<<CAIXA_LIMPEZA_6M>>'},
    'limp_caixa_mais_6m':{'col':'CAIXA_LIMP', 'text':'Sim, há mais de 6 meses', 'tag':'<<CAIXA_LIMPEZA_MAIS_6M>>'},
    'limp_caixa_nao':{'col':'CAIXA_LIMP', 'text':'Não', 'tag':'<<CAIXA_LIMPEZA_NAO>>'},

    'limp_cisterna_6m':{'col':'CIST_LIMP', 'text':'Sim, nos últimos 6 meses', 'tag':'<<CISTERNA_LIMPEZA_6M>>'},
    'limp_cisterna_mais_6m':{'col':'CIST_LIMP', 'text':'Sim, há mais de 6 meses', 'tag':'<<CISTERNA_LIMPEZA_MAIS_6M>>'},
    'limp_cisterna_nao':{'col':'CIST_LIMP', 'text':'Não', 'tag':'<<CISTERNA_LIMPEZA_NAO>>'},

    'agua_certif_sim':{'col':'AGUA_CERTIF', 'text':'Sim', 'tag':'<<AGUA_CERTIFICADO_SIM>>'},
    'agua_certif_nao':{'col':'AGUA_CERTIF', 'text':'Não', 'tag':'<<AGUA_CERTIFICADO_NAO>>'},
    
    'agua_potavel_sim':{'col':'AGUA_POT', 'text':'Sim', 'tag':'<<AGUA_POTAVEL_DIA_SIM>>'},
    'agua_potavel_nao':{'col':'AGUA_POT', 'text':'Não', 'tag':'<<AGUA_POTAVEL_DIA_NAO>>'},

    'abastecimento_desconformidades_sim':{'col':'ABAST_DESC', 'text':'Sim', 'tag':'<<AGUA_ABAST_DESC_SIM>>'},
    'abastecimento_desconformidades_nao':{'col':'ABAST_DESC', 'text':'Não', 'tag':'<<AGUA_ABAST_DESC_NAO>>'},

    'risco_grave':{'col':'RISCO_GRAVE', 'op':'composite','conditions':{'must_contain':["Sim"]}, 'tag':'<<RISCO_GRAVE_SIM>>'},
    'risco_grave_nao':{'col':'RISCO_GRAVE', 'text':'Não', 'tag':'<<RISCO_GRAVE_NAO>>'},

    # Perguntas Direcionadas à merendeira
    'recebimento_adequado':{'col':'REC_ADQ', 'text':'Sim', 'tag':'<<OFERTA_RECEBIMENTO_ADEQUADO_SIM>>'},
    'recebimento_adequado_nao':{'col':'REC_ADQ', 'text':'Não', 'tag':'<<OFERTA_RECEBIMENTO_ADEQUADO_NAO>>'},

    'quantidade_adequada':{'col':'QTD_ADQ', 'text':'Sim, a quantidade é adequada', 'tag':'<<OFERTA_QTD_ADEQUADA>>'},
    'quantidade_adequada_sobras':{'col':'QTD_ADQ', 'text':'Sim, havendo ocorrência de sobras', 'tag':'<<OFERTA_QTD_SIM_SOBRAS>>'},
    'quantidade_nao_adequada':{'col':'QTD_ADQ', 'text':'Não', 'tag':'<<OFERTA_QTD_NAO>>'},

    'falta_merenda':{'col':'FALTA_MERENDA', 'text':'Sim', 'tag':'<<OFERTA_FALTA_MERENDA_SIM>>'},
    'nao_falta_merenda':{'col':'FALTA_MERENDA', 'text':'Não', 'tag':'<<OFERTA_FALTA_MERENDA_NAO>>'},

    # se der trabalho, origem de água vem aq tbm
    # ... etc
}

# Especificações para tags "dinâmicas" (que extraem palavras da própria tag).
# Formato: 'PREFIXO_DA_TAG': {'col': 'NOME_DA_COLUNA', 'replacements': {'DE': 'PARA'}}
DYNAMIC_TAG_SPECS = {
    'DEP_': {'col': 'DEPENDENCIAS', 'replacements': {"_": " "}},
    #TODO colocar DEP_PATIO como tag estática. 
    'DEP_INF_': {'col': 'DEP_INF', 'replacements': {"_": " "}},
    'NUM_ESCOLAS_': {'col': 'ACESSO', 'replacements': {"_": " ", "INADEQ":""}},
    'COZ_INFRA_': {'col': 'IRR_COZ', 'replacements': {"VENT": "VENTILACAO", "ILUM": "ILUMINACAO", "_": " "}},
    'LIXO_':{'col':'LIXO', 'replacements':{"_":" "}},
    'ARMZ_UP':{'col':'ALM_UP', 'replacements':{"_":" ","ARMZ_UP":""}},
    'UP_CONSUMO':{'col':'CONS_UP', 'replacements':{"_":" ","UP_CONSUMO":""}},
    'UP_DIRETRIZ':{'col':'DIR_UP', 'replacements':{"_":" ","UP_DIRETRIZ":""}},
    'UP_COMERCIO':{'col':'VEND_UP', 'replacements':{"_":" ","UP_COMERCIO":""}},
    'CARDAPIO_CONFORME':{'col':'REF_CONF', 'replacements':{"_":" ","CARDAPIO_CONFORME":""}},
    'COZ_OUTROS':{'col':'COZ_OUT', 'replacements':{"_":" ","COZ_OUTROS":""}},
    'FRALDARIO_ITEM':{'col':'ITEM_FRAL', 'replacements':{"_":" ","FRALDARIO_ITEM":""}},
    'FRALDARIO_INFRA':{'col':'IRR_FRAL', 'replacements':{"_":" ","FRALDARIO_INFRA":"","VENT":"VENTILAÇÃO", "ILUM":"ILUMINAÇÃO", "PROBLEMA": ""}},
    'LACTARIO_LOCAL':{'col':'LOCAL_LACT', 'replacements':{"_":" ","LACTARIO_LOCAL":"","COZINHA":""}},
    'LACTARIO_ITEM':{'col':'ITEM_LACT', 'replacements':{"_":" ","LACTARIO_ITEM":""}},
    'LACTARIO_INFRA':{'col':'IRR_LACT','replacements':{"_":" ","LACTARIO_INFRA":"","AUS":"AUSENCIA","VENT":"VENTILACAO"}},
    'BERCARIO_INFRA':{'col':'IRR_BER','replacements':{"_":" ","BERCARIO_INFRA":"","AUS":"AUSENCIA","VENT_NAT":"VENTILACAO_NATURAL","VENT_MEC":"VENTILADORES"}},
    'ARMZ_CONGELADOS':{'col':'ALM_CONG','replacements':{"_":" ","ARMZ_CONGELADOS":""}},
    'ARMZ_CONG_IRREG':{'col':'IRR_CONG','replacements':{"_":" ","ARMZ_CONG_IRREG":"","ENCONTRADO":""}},
    'EI_SAN_EXC_AGUA':{'col':'AG_SAN_EI','replacements':{"_":" ","EI_SAN_EXC_AGUA":""}},
    'EI_SAN_EXC_INFRA':{'col':'IRR_SAN_EI','replacements':{"_":" ","EI_SAN_EXC_INFRA":"","NENHUMA":"NAO","PROBLEMA":""}},
    'BIBLI_SL_COMP_INFRA':{'col':'IRR_BIBSL','replacements':{"_":" ","BIBLI_SL_COMP_INFRA":"","AUS":"AUSENCIA","NENHUMA":"ASPECTOS","IRREGULARES":"","IRREGULAR":"",}},
    'BANH_GERAL_INFRA':{'col':'IRR_BAN','replacements':{"_":" ","BANH_GERAL_INFRA":"","PROBLEMA":"","AUS":"AUSENCIA","ILUM":"ILUMINACAO","VENT":"VENTILACAO"}},
    'REFEITORIO_MOB':{'col':'MOB_REF','replacements':{"REFEITORIO_MOB":"","OK":"NAO_FORAM","SEM_ADAPT":"AUSENCIA_ADAPTACAO","_":" "}},
    'REFEITORIO_INFRA':{'col':'IRR_REF','replacements':{"REFEITORIO_INFRA":"","ILUM":"ILUMINACAO","VENT":"VENTILACAO","OK":"NAO_FORAM","_":" "}},
    'RESERVATORIO':{'col':'RES_AGUA','replacements':{"RESERVATORIO":"","FUNC":"FUNCIONAMENTO","OK":"EM_FUNCIONAMENTO","_":" "}},
    'PCD_SAN_EXC_AGUA':{'col':'BAN_PCD_AG','replacements':{"PCD_SAN_EXC_AGUA":"","AGUA":"","_":" "}},
    'PCD_SAN_INFRA':{'col':'IRR_BAN_PCD','replacements':{"PCD_SAN_INFRA":"","AUS":"","NENHUMA_IRREG":"NAO_FORAM","_":" "}},
    'BIB_EXC_INFRA':{'col':'IRR_BIB','replacements':{"BIB_EXC_INFRA":"","_":"","AUS":"AUSÊNCIA","NENHUMA":"ASPECTOS","IRREGULARES":"","IRREGULAR":""}},
    'SL_EXC_INFRA':{'col':'IRR_SL','replacements':{"SL_EXC_INFRA":"","_":"","AUS":"AUSÊNCIA","NENHUMA":"ASPECTOS","IRREGULARES":"","IRREGULAR":""}},
    'PCD_SAN_ACESS':{'col':'ITEM_PCD','replacements':{"PCD_SAN_ACESS":"","_":"","ESPECIAL":"","VAO":"","ANTIDERRAPANTE":"","APOIO":"","80":"","PORTA":""}},
    'AGUA_ORIGEM':{'col':'ORIGEM_AGUA', 'replacements':{"AGUA_ORIGEM":"","_":"","REDE":"","NAO":"","FILTRO":"","FONTE":""}},
    'OFERTA_FRUTAS':{'col':'OFERTA_FRUTA', 'replacements':{"OFERTA_FRUTAS":"","_D":""}},
    'OFERTA_LEGUMES':{'col':'OFERTA_LEGUME', 'replacements':{"OFERTA_LEGUMES":"","_D":""}},



    # Adicionar outras lógicas dinâmicas aqui
    # O script ainda detecta LACTARIO_LOCAL? Vê se o _ interfere (interfere :p)
}


# ========================= 2. FUNÇÕES AUXILIARES E DE PROCESSAMENTO =========================

def _col_letter_to_index(letra: str) -> int:
    """Converte letra de coluna Excel para índice numérico (A=0, B=1, ...)."""
    result = 0
    for char in letra.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1

def _carregar_tags(caminho: Path) -> List[str]:
    """Carrega lista de tags do arquivo texto."""
    print(f"Carregando tags de: {caminho}")
    with open(caminho, 'r', encoding='utf-8') as f:
        tags = [linha.strip() for linha in f if linha.strip()]
    print(f"{len(tags)} tags carregadas.")
    return tags

def _norm_noacc(s: str) -> str:
    """Normaliza string: remove acentos, múltiplos espaços e converte para minúsculas."""
    s = str(s or "")
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).lower().strip()

def carregar_e_preparar_base(caminho_bd: Path) -> pd.DataFrame:
    """Carrega a planilha e renomeia as colunas para nomes amigáveis."""
    print(f"Carregando base de dados de: {caminho_bd}")
    df = pd.read_excel(caminho_bd, sheet_name=0, header=0)
    print(f"{len(df)} linhas carregadas.")

    # Mapeia letras das colunas para os nomes definidos em COLS_AS
    col_indices = {_col_letter_to_index(v): k for k, v in COLS_AS.items()}
    
    # Renomeia as colunas que serão usadas
    df.rename(columns={df.columns[i]: name for i, name in col_indices.items() if i < len(df.columns)}, inplace=True)

    # Normaliza colunas de texto para busca mais rápida e consistente
    for col_name in COLS_AS.keys():
        if col_name in df.columns and df[col_name].dtype == 'object':
            df[f"{col_name}_norm"] = df[col_name].apply(_norm_noacc)
            
    print("Pré-processamento da base concluído.")
    return df

def build_tag_registry(specs: Dict[str, Dict]) -> Dict[str, Dict]:
    """Constrói o registro completo de tags, incluindo variações _MUN e _EST."""
    registry = {}
    for key, spec in specs.items():
        base_tag = spec['tag']
        # Tag Geral
        registry[base_tag] = {'spec_key': key, 'esfera': None}
        # Tag Municipal
        registry[base_tag.replace(">>", "_MUN>>")] = {'spec_key': key, 'esfera': 'municipal'}
        # Tag Estadual
        registry[base_tag.replace(">>", "_EST>>")] = {'spec_key': key, 'esfera': 'estadual'}
    return registry


# ========================= 3. O NOVO MOTOR DE CÁLCULO =========================

def calcular_tags_para_municipio(
    municipio: str, 
    df: pd.DataFrame, 
    tags_a_calcular: List[str],
    static_registry: Dict[str, Dict],
    dynamic_specs: Dict[str, Dict]
) -> Dict[str, str]:
    """
    Calcula todas as tags para um município específico de forma vetorizada.
    """
    print(f"  Processando município: {municipio}")
    resultados = {}
    
    # Filtro base para o município (feito uma única vez)
    df_municipio = df[df['MUNICIPIO'] == municipio].copy()
    if df_municipio.empty:
        print(f"    Aviso: Nenhum dado encontrado para o município '{municipio}'.")
        return {tag: "0" for tag in tags_a_calcular}

    # Cache para cálculos já feitos (evita reprocessar a mesma condição)
    calc_cache = {}

    for tag in tags_a_calcular:
        # --- Lógica para Tags Estáticas ---
        if tag in static_registry:
            rule = static_registry[tag]
            spec_key = rule['spec_key']
            
            if spec_key not in calc_cache:
                spec = CALCULATION_SPECS[spec_key]
                col, op, text = spec.get('col'), spec.get('op'), spec.get('text')
                if op == 'notna':
                    mask = df_municipio[col].notna()

                elif op == 'composite':
                    # Começamos com uma máscara que inclui todas as linhas
                    base_mask = pd.Series([True] * len(df_municipio), index=df_municipio.index)
                    col_norm = f"{col}_norm"
                    conditions = spec.get('conditions', {})

                    # Aplica as condições 'must_contain' (Lógica AND)
                    for text_to_find in conditions.get('must_contain', []):
                        base_mask &= df_municipio[col_norm].str.contains(_norm_noacc(text_to_find), na=False)
                    
                    # Aplica as condições 'must_not_contain' (Lógica AND NOT)
                    for text_to_exclude in conditions.get('must_not_contain', []):
                        # O operador ~ inverte a máscara booleana (faz um NOT)
                        base_mask &= ~df_municipio[col_norm].str.match(_norm_noacc(text_to_exclude), na=False)
                    
                    mask = base_mask

                else: # 'contains' é o padrão
                    mask = df_municipio[f"{col}_norm"].str.match(_norm_noacc(text), na=False)
                calc_cache[spec_key] = mask

            final_mask = calc_cache[spec_key].copy()
            if rule['esfera']:
                # ===== LINHA CORRIGIDA AQUI =====
                final_mask &= df_municipio['ESFERA_norm'].str.contains(rule['esfera'].lower(), na=False)
                
            resultados[tag] = str(final_mask.sum())
            continue

        # --- Lógica para Tags Dinâmicas ---
        found_dynamic = False
        for prefix, spec in dynamic_specs.items():
            if tag.startswith(f"<<{prefix}"):
                col_name = spec['col']
    
                sufixo = tag.replace("<<", "").replace(">>", "").replace(prefix, "")
                esfera = None

                if sufixo.endswith("_MUN"):
                    esfera = "Municipal"
                    sufixo = sufixo[:-4]
                elif sufixo.endswith("_EST"):
                    esfera = "Estadual"
                    sufixo = sufixo[:-4]

                for de, para in spec.get('replacements', {}).items():
                    sufixo = sufixo.replace(de, para)

                tokens = [t for t in sufixo.split(" ") if t]
                
                if not tokens:
                    combined_mask = pd.Series([True] * len(df_municipio), index=df_municipio.index)
                else:
                    masks = [df_municipio[f"{col_name}_norm"].str.contains(_norm_noacc(token), na=False) for token in tokens]
                    combined_mask = reduce(operator.and_, masks)

                if esfera:
                    # ===== LINHA CORRIGIDA AQUI =====
                    combined_mask &= df_municipio['ESFERA_norm'].str.contains(esfera.lower(), na=False)

                resultados[tag] = str(combined_mask.sum())
                found_dynamic = True
                break

        if not found_dynamic and tag not in static_registry:
            resultados[tag] = " "

    for tag in tags_a_calcular:
        if tag not in resultados:
            resultados[tag] = " "
            
    return resultados

# ========================= 4. FUNÇÃO DE ESCRITA E MAIN (praticamente inalteradas) =========================

def _escrever_planilha(saida: Path, tags: List[str], municipios: List[str], resultados_por_municipio: Dict[str, Dict[str, str]]):
    """Escreve a planilha de saída com os resultados."""
    print(f"Gerando planilha: {saida}")
    wb = Workbook()
    ws_calc = wb.active
    ws_calc.title = "ListaTagsCalculadas"

    # Cabeçalhos
    ws_calc["A1"] = "Município Selecionado:"
    ws_calc["B1"] = "=ListaTagsSimples!B1"

    ws_calc["A2"] = "Tags a serem substituídas"
    ws_calc["B2"] = ""
    ws_calc["C2"] = "Script"
    ws_calc["D2"] = "Descrição"

    start_col_data = 5 # Coluna B para os dados

    # Cabeçalho de municípios (Linha 3)
    for j, m in enumerate(municipios, start=start_col_data):
        ws_calc.cell(row=2, column=j, value=m)
    
      # Escreve tags e resultados
    for i, tag in enumerate(tags, start=4):
        ws_calc.cell(row=i, column=1, value=tag) # Coluna A
        for j, municipio in enumerate(municipios, start=start_col_data):
            valor = resultados_por_municipio.get(municipio, {}).get(tag, "")
            ws_calc.cell(row=i, column=j, value=valor)

    # Fórmula B3 (igual à sua)
    if tags:
        ws_calc["B3"] = (
            "=LET("
            "mun, ListaTagsSimples!B1,"
            "headers, E2:2,"
            "matriz, E3:AAA,"
            "col, CORRESP(mun, headers, 0),"
            "ÍNDICE(matriz, , col)"
            ")"
        )

    # Ajuste de larguras
    ws_calc.column_dimensions["A"].width = 50
    for j in range(start_col_data, start_col_data + len(municipios)):
        ws_calc.column_dimensions[get_column_letter(j)].width = 22

    wb.save(saida)
    print("Planilha salva com sucesso!")

def main():
    ap = argparse.ArgumentParser(description="Calcula tags de forma otimizada a partir de uma base de dados.")
    ap.add_argument("--basedados", required=True, type=Path, help="Caminho para a planilha de base de dados (ex: bd.xlsx)")
    ap.add_argument("--tags", required=True, type=Path, help="Caminho para o arquivo de texto com as tags (ex: TAGS_a_calcular.txt)")
    ap.add_argument("--saida", required=True, type=Path, help="Caminho para a planilha de saída (ex: TagsCalculadas.xlsx)")
    ap.add_argument("--municipio", help="(Opcional) Nome de um município específico para calcular. Se não for fornecido, calcula para todos.")
    args = ap.parse_args()

    # 1. Carregar e preparar dados
    df = carregar_e_preparar_base(args.basedados)
    tags_a_calcular = _carregar_tags(args.tags)
    
    # 2. Construir o registro de tags
    static_registry = build_tag_registry(CALCULATION_SPECS)

    # 3. Determinar para quais municípios calcular
    if args.municipio:
        municipios_a_processar = [args.municipio]
    else:
        municipios_a_processar = sorted(df['MUNICIPIO'].dropna().unique())
        
        print(f"Nenhum município especificado. Calculando para todos os {len(municipios_a_processar)} municípios encontrados.")

    # 4. Calcular para cada município
    resultados_gerais = {}
    for municipio in municipios_a_processar:
        resultados_gerais[municipio] = calcular_tags_para_municipio(
            municipio, df, tags_a_calcular, static_registry, DYNAMIC_TAG_SPECS
        )

    # 5. Escrever a planilha de saída
    _escrever_planilha(args.saida, tags_a_calcular, municipios_a_processar, resultados_gerais)
    
    print("\nProcessamento concluído com sucesso!")
    print(f"Arquivo gerado: {args.saida}")

if __name__ == "__main__":
    main()
