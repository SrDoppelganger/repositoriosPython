#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script unificado para calcular todas as tags (escolas visitadas + ANVISA/vigilância sanitária)
"""

import argparse
import re
import unicodedata
import pandas as pd
from pathlib import Path
from typing import Dict, List, Set
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ========================= CONFIGURAÇÕES =========================
COLS_AS = {
    "MUNICIPIO": "F",
    "INEP": "G",
    "ESFERA": "K",
    "ACESSO":"Z",
    "ETAPAS": "AJ",
    "ESGOTAMENTO":"EE",
    "DEPENDENCIAS": "ER",
    "DEP_INF":"ET",

    "IRR_ENTRADA": "AE", # Coluna de presença de rampas
    "GEST_ALIM": "BL", # Coluna de gestão de alimentação
    "TERC_ALIM":"BN", #Coluna de tipo de terceirização de alimentação
    "ANVISA": "CK",  # Coluna de vigilância sanitária
    "AVCB": "CP", # Coluna de vigilância sanitária
    "DEDET": "CU", # Coluna de vigilância sanitária
    "ABS_AGUA":"CZ", # Coluna sobre abastecimento água
    "RES_AGUA":"DB", # Coluna sobre reservatório de água
    "LIXO":"EK", # Coluna sobre destinação do lixo
    "ENERGIA":"EP", #Coluna sobre conexão com a rede de energia elétrica
    "PATIO":"EV", # Coluna sobre patio exclusivo/compart.
    "PARQ_INF":"FA", # Coluna sobre brinq. no parquinho infantil
    "IRR_BIB":"FV", # Coluna sobre irregularidades nas bibliotecas (espaço exclusivo)
    "IRR_SL":"GA", # Coluna sobre irregularidades na Sala de Leitura (espaço exclusivo)
    "IRR_BIBSL":"GF", # Coluna sobre irregularidades em Sala de Leitura + Biblioteca (espaço compartilhado)
    "AG_SAN_EI":"GN", # Coluna sobre agua nos banheiros da educ infantil
    "IRR_SAN_EI":"GV", # Coluna sobre irregularidades nos banheiros da educ infantil
    "MULTS":"II", # Coluna sobre salas multisseriadas
    "SALA_IRR":"IU", # Coluna sobre irregularidades nas salas de aula
    "ITEM_LACT":"KE", # Coluna sobre itens nos lactários
    "IRR_LACT":"KJ", # Coluna sobre irregularidades nos lactários
    "LOCAL_FRAL":"KO", # Coluna sobre localidade dos fraldários
    "ITEM_FRAL":"KT", # Coluna sobre itens nos fraldários
    "IRR_FRAL":"KV", # Coluna sobre irregularidades nos fraldários
    "EQP_COZ":"LF", # Coluna sobre irregularidades nos fraldários
    "COZ_OUT":"LR", # Coluna sobre outras irregularidades na cozinha
    "LOC_ARM":"LW", # Coluna do local de armazenamento
    "LIM_BERC":"JD", # Coluna sobre o limite de crianças no berçário
    "IRR_BER":"JL", # Coluna sobre irregularidaes nos berçários
    "IRR_COZ":"LM", # Coluna sobre irregularidades nas cozinhas
    "ARM_IRR":"MB", # Coluna sobre irregularidades nos locais de armazenamento
    "ALM_IRR":"MG", # Coluna sobre irregularidades nos alimentos armazenados
    "ALM_UP":"ML", # Coluna sobre limentos ultraprocessados
    "ALM_CONG":"MV", # Coluna sobre armazenamento de alimentos congelados
    "IRR_CONG":"NA", # Coluna sobre irregularidades no armazenamento de alimentos congelados
    "CARD":"NW", # Coluna sobre o cardápio
    "CARD_ESP":"OB", # Coluna sobre cardápio especial
    "REF_SERV":"OG", # Coluna sobre refeição sendo servida na hora da visita
    "CARD_CONF": "OM", # Coluna sobre refeições de acordo com cardápio
    "IRR_REF":"PB", # Coluna sobre irregularidade dnos refeitórios
    "CONS_UP":"PG", # Coluna sobre consumo de alimentos ultraprocessados pelos alunos
    "VEND_UP":"PL", # Coluna sobre a venda de alimentos ultraprocessados
    "DIR_UP":"PR", # Coluna sobre diretriz de ultraprocessados
}

# ========================= FUNÇÕES AUXILIARES =========================
def _col_letter_to_index(letra: str) -> int:
    """Converte letra de coluna Excel para índice numérico (A=0, B=1, ...)."""
    result = 0
    for char in letra.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1

def _carregar_base_dados(caminho: Path) -> pd.DataFrame:
    """Carrega a planilha de dados."""
    print(f"Carregando base de dados de: {caminho}")
    df = pd.read_excel(caminho, sheet_name=0)
    print(f"{len(df)} linhas carregadas")
    return df

def _carregar_tags(caminho: Path) -> List[str]:
    """Carrega lista de tags do arquivo texto."""
    print(f"Carregando tags de: {caminho}")
    with open(caminho, 'r', encoding='utf-8') as f:
        tags = [linha.strip() for linha in f if linha.strip()]
    print(f"{len(tags)} tags carregadas")
    return tags

def _extrair_municipios(df: pd.DataFrame) -> List[str]:
    """Extrai lista única de municípios da planilha."""
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    municipios = df.iloc[:, i_mun].dropna().unique().tolist()
    return sorted([str(m) for m in municipios])

def _norm_noacc(s: str) -> str:
    s = (s or "")
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = s.lower().strip()
    s = re.sub(r"\s+", " ", s)
    return s

# ========================= CÁLCULO: ESCOLAS VISITADAS =========================
def _calc_escolas_visitadas(municipio: str, df: pd.DataFrame) -> Dict[str, Set[str]]:
    """Port direto do Apps Script para ESCOLAS (usando letras de coluna)."""
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_etp = _col_letter_to_index(COLS_AS["ETAPAS"])

    todas, municipais, estaduais = set(), set(), set()
    inf_creche, inf_pre = set(), set()
    inf_q, inf_q_mun, inf_q_est = set(), set(), set()
    inf_creche_mun, inf_creche_est = set(), set()
    inf_pre_mun, inf_pre_est = set(), set()
    fund_q, fund_q_mun, fund_q_est = set(), set(), set()
    med_q, med_q_mun, med_q_est = set(), set(), set()
    eja_q, eja_q_mun, eja_q_est = set(), set(), set()
    inf_fund_exclusivo, inf_fund_exclusivo_mun, inf_fund_exclusivo_est = set(), set(), set()
    fund_exclusivo, med_exclusivo = set(), set()


    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        esfera = str(row.iloc[i_esf]).strip()
        etapas = str(row.iloc[i_etp])

        todas.add(inep)
        if esfera == "Municipal":
            municipais.add(inep)
        elif esfera == "Estadual":
            estaduais.add(inep)

        tem_creche = "Educação Infantil - Creche" in etapas
        tem_pre    = "Educação Infantil - Pré-escola" in etapas
        tem_inf    = tem_creche or tem_pre
        tem_fund   = ("Fundamental - Anos iniciais" in etapas) or ("Fundamental - Anos finais" in etapas)
        tem_med    = "Ensino Médio" in etapas
        tem_eja    = "EJA" in etapas

        # exclusivos
        if tem_inf and tem_fund and not tem_med:
            inf_fund_exclusivo.add(inep)
            if esfera == "Municipal":
                inf_fund_exclusivo_mun.add(inep)
            elif esfera == "Estadual":
                inf_fund_exclusivo_est.add(inep)
        if tem_fund and not tem_inf and not tem_med:
            fund_exclusivo.add(inep)
        if tem_med and not tem_fund and not tem_inf:
            med_exclusivo.add(inep)

        if tem_creche:
            inf_creche.add(inep)
            if esfera == "Municipal": inf_creche_mun.add(inep)
            elif esfera == "Estadual": inf_creche_est.add(inep)
        if tem_pre:
            inf_pre.add(inep)
            if esfera == "Municipal": inf_pre_mun.add(inep)
            elif esfera == "Estadual": inf_pre_est.add(inep)

        if tem_inf:
            inf_q.add(inep)
            if esfera == "Municipal": inf_q_mun.add(inep)
            elif esfera == "Estadual": inf_q_est.add(inep)

        if tem_fund:
            fund_q.add(inep)
            if esfera == "Municipal": fund_q_mun.add(inep)
            elif esfera == "Estadual": fund_q_est.add(inep)

        if tem_med:
            med_q.add(inep)
            if esfera == "Municipal": med_q_mun.add(inep)
            elif esfera == "Estadual": med_q_est.add(inep)

        if tem_eja:
            eja_q.add(inep)
            if esfera == "Municipal": eja_q_mun.add(inep)
            elif esfera == "Estadual": eja_q_est.add(inep)

    return {
        "todas": todas,
        "municipais": municipais,
        "estaduais": estaduais,

        "infantil_q": inf_q,
        "infantil_q_mun": inf_q_mun,
        "infantil_q_est": inf_q_est,

        "infantil_creche": inf_creche,
        "inf_creche_mun": inf_creche_mun,
        "inf_creche_est": inf_creche_est,

        "infantil_pre": inf_pre,
        "inf_pre_mun": inf_pre_mun,
        "inf_pre_est": inf_pre_est,

        "fund_q": fund_q,
        "fund_q_mun": fund_q_mun,
        "fund_q_est": fund_q_est,

        "medio_q": med_q,
        "medio_q_mun": med_q_mun,
        "medio_q_est": med_q_est,

        "eja_q": eja_q,
        "eja_q_mun": eja_q_mun,
        "eja_q_est": eja_q_est,

        "inf_fund_exclusivo": inf_fund_exclusivo,
        "inf_fund_exclusivo_mun": inf_fund_exclusivo_mun,
        "inf_fund_exclusivo_est": inf_fund_exclusivo_est,

        "fund_exclusivo": fund_exclusivo,
        "med_exclusivo": med_exclusivo
    }
# ========================= CÁLCULO: PATIO/AREA PARA ATV EXTERNAS =========================
def _calc_infra_patio(municipio: str, df: pd.DataFrame) -> Dict[str, Set[str]]:
    """Port direto do Apps Script para ESCOLAS (usando letras de coluna)."""
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_etp = _col_letter_to_index(COLS_AS["PATIO"])

    #EXCLUSIVO
    cob_ex, cob_ex_mun, cob_ex_est = set(), set(), set()
    descob_ex, descob_ex_mun, descob_ex_est = set(), set(), set()
    #COMPARTILHADO
    cob_comp, cob_comp_mun, cob_comp_est = set(), set(), set()
    descob_comp, descob_comp_mun, descob_comp_est = set(), set(), set()
    #TOTAL
    nao, nao_mun, nao_est = set(), set(), set()
   


    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        esfera = str(row.iloc[i_esf]).strip()
        patios = str(row.iloc[i_etp])


        tem_cob_ex = "Sim, área COBERTA com horário de utilização EXCLUSIVO para o ensino infantil" in patios
        tem_descob_ex = "Sim, área DESCOBERTA com horário de utilização EXCLUSIVO para o ensino infantil" in patios
        tem_cob_comp = "Sim, área COBERTA COMPARTILHADA no mesmo horário com outras etapas de ensino" in patios
        tem_descob_comp = "Sim, área DESCOBERTA COMPARTILHADA no mesmo horário com outras etapas de ensino" in patios
        nao_tem = "Não" in patios

        if nao_tem:
            nao.add(inep)
            if esfera == "Municipal": nao_mun.add(inep)
            elif esfera == "Estadual": nao_est.add(inep)

        if tem_cob_ex:
            cob_ex.add(inep)
            if esfera == "Municipal": cob_ex_mun.add(inep)
            elif esfera == "Estadual": cob_ex_est.add(inep)
        if tem_descob_ex:
            descob_ex.add(inep)
            if esfera == "Municipal": descob_ex_mun.add(inep)
            elif esfera == "Estadual": descob_ex_est.add(inep)

        if tem_cob_comp:
            cob_comp.add(inep)
            if esfera == "Municipal": cob_comp_mun.add(inep)
            elif esfera == "Estadual": cob_comp_est.add(inep)
        if tem_descob_comp:
            descob_comp.add(inep)
            if esfera == "Municipal": descob_comp_mun.add(inep)
            elif esfera == "Estadual": descob_comp_est.add(inep)

    return {
        "nao": nao,
        "nao_mun": nao_mun,
        "nao_est": nao_est,

        "cob_ex": cob_ex,
        "cob_ex_mun": cob_ex_mun,
        "cob_ex_est": cob_ex_est,

        "descob_ex": descob_ex,
        "descob_ex_mun": descob_ex_mun,
        "descob_ex_est": descob_ex_est,

        "cob_comp": cob_comp,
        "cob_comp_mun": cob_comp_mun,
        "cob_comp_est": cob_comp_est,

        "descob_comp": descob_comp,
        "descob_comp_mun": descob_comp_mun,
        "descob_comp_est": descob_comp_est,
    }
# ========================= CÁLCULO: DEPS infantil  =========================
def _deps_inf_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("DEP", "").replace("<<", "").replace(">>", "")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _deps_inf_match(er_norm: str, acesso_key: str) -> bool:
    toks = _deps_inf_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_deps_inf_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["DEP_INF"])
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _deps_inf_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: GESTÃO DA ALIMENTAÇÃO  =========================
def _calc_gestao_alim(municipio: str, df: pd.DataFrame) -> Dict[dict, Set[str]]:
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_tag= _col_letter_to_index(COLS_AS["GEST_ALIM"])

    cent, cent_mun, cent_est = set(), set(), set()
    descent, descent_mun, descent_est = set(), set(), set()
    semicent, semicent_mun, semicent_est = set(), set(), set()
    terc, terc_mun, terc_est = set(), set(), set()

    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        resp = str(row.iloc[i_tag])

        centralizada = "Centralizada" in resp
        escolarizada = "Descentralizada ou Escolarizada" in resp
        semi_escolarizada = "Semi Descentralizada ou Parcialmente Escolarizada" in resp
        terceirizada = "Terceirizada" in resp

        if centralizada:
            cent.add(inep)
            if esfera == "Municipal": cent_mun.add(inep)
            elif esfera == "Estadual": cent_est.add(inep)
        if escolarizada:
            descent.add(inep)
            if esfera == "Municipal": descent_mun.add(inep)
            elif esfera == "Estadual": descent_est.add(inep)
        if semi_escolarizada:
            semicent.add(inep)
            if esfera == "Municipal": semicent_mun.add(inep)
            elif esfera == "Estadual": semicent_est.add(inep)
        if terceirizada:
            terc.add(inep)
            if esfera == "Municipal": terc_mun.add(inep)
            elif esfera == "Estadual": terc_est.add(inep)
 
    return {
        "centralizada": cent,
        "centralizada_mun": cent_mun,
        "centralizada_est": cent_est,

        "descentralizada": descent,
        "descentralizada_mun": descent_mun,
        "descentralizada_est": descent_est,

        "semidescentralizada": semicent,
        "semidescentralizada_mun": semicent_mun,
        "semidescentralizada_est": semicent_est,

        "terceirizada": terc,
        "terceirizada_mun": terc_mun,
        "terceirizada_est": terc_est,
    }

# ========================= CÁLCULO: (ALIMENTAÇÃO) TIPO DE TERCEIRIZAÇÃO =========================
def _calc_terceirizacao_alim(municipio: str, df: pd.DataFrame) -> Dict[dict, Set[str]]:
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_arm = _col_letter_to_index(COLS_AS["TERC_ALIM"])

    eqp_escola, eqp_escola_mun, eqp_escola_est = set(), set(), set()
    eqp_empresa, eqp_empresa_mun, eqp_empresa_est = set(), set(), set()
    hotbox, hotbox_mun, hotbox_est = set(), set(), set()

    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        tipo = str(row.iloc[i_arm])

        prep_escola = "Com preparo na escola, usando equipamentos da escola" in tipo
        prep_empresa = "Com preparo na escola, usando equipamentos da empresa" in tipo
        entrega = "Com entrega em hot box" in tipo

        if prep_escola:
            eqp_escola.add(inep)
            if esfera == "Municipal": eqp_escola_mun.add(inep)
            elif esfera == "Estadual": eqp_escola_est.add(inep)
        if prep_empresa:
            eqp_empresa.add(inep)
            if esfera == "Municipal": eqp_empresa_mun.add(inep)
            elif esfera == "Estadual": eqp_empresa_est.add(inep)
        if entrega:
            hotbox.add(inep)
            if esfera == "Municipal": hotbox_mun.add(inep)
            elif esfera == "Estadual": hotbox_est.add(inep)

    return {
        "eqp_escola": eqp_escola,
        "eqp_escola_mun": eqp_escola_mun,
        "eqp_escola_est": eqp_escola_est,

        "eqp_empresa": eqp_empresa,
        "eqp_empresa_mun": eqp_empresa_mun,
        "eqp_empresa_est": eqp_empresa_est,

        "hotbox": hotbox,
        "hotbox_mun": hotbox_mun,
        "hotbox_est": hotbox_est,
    }

# ========================= CÁLCULO: INADEQUAÇÕES NA SEGURANÇA E ACESSO =========================
def _acesso_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("NUM_ESCOLAS", "").replace("INADEQ", "").replace("<<", "").replace(">>", "")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _acesso_match(er_norm: str, acesso_key: str) -> bool:
    toks = _acesso_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_acessibilidade_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["ACESSO"])
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _acesso_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: EQUIPAMENTOS COZINHA =========================

# FAZER LÓGICA PARA VERIFICAR SE TEM TODAS, ALGUMAS OU NENHUMA FUNCIONANDO

# ========================= CÁLCULO: INFRAESTRUTURA COZINHA =========================
def _infra_cozinha_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("COZ_INFRA", "").replace("VENT", "VENTILACAO").replace("ILUM","ILUMINACAO").replace("<<", "").replace(">>", "")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _infra_cozinha_match(er_norm: str, acesso_key: str) -> bool:
    toks = _infra_cozinha_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_infra_cozinha_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["IRR_COZ"])
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _infra_cozinha_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: COZINHA (OUTROS) =========================
def _cozinha_outros_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("COZ_OUTROS", "").replace("<<", "").replace(">>", "")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _cozinha_outros_match(er_norm: str, acesso_key: str) -> bool:
    toks = _cozinha_outros_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_cozinha_outros_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["COZ_OUT"])
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _cozinha_outros_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: LOCAL DE ARMAZENAMENTO DE GEN. ALIMENTÍCIOS  =========================
def _calc_armazenamento_alim(municipio: str, df: pd.DataFrame) -> Dict[dict, Set[str]]:
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_arm = _col_letter_to_index(COLS_AS["LOC_ARM"])

    despensa, despensa_mun, despensa_est = set(), set(), set()
    arm_dentro, arm_dentro_mun, arm_dentro_est = set(), set(), set()
    arm_fora, arm_fora_mun, arm_fora_est = set(), set(), set()
    nao_ha, nao_ha_mun, nao_ha_est = set(), set(), set()

    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        armazenamento = str(row.iloc[i_arm])

        em_despensa = "Em local especifico de despensa" in armazenamento
        dentro = "Em armário, dentro na cozinha" in armazenamento
        fora = "Em armário, fora do ambiente da cozinha" in armazenamento
        sem_arm = "Não há local de armazenamento de gêneros alimentícios na escola" in armazenamento

        if em_despensa:
            despensa.add(inep)
            if esfera == "Municipal": despensa_mun.add(inep)
            elif esfera == "Estadual": despensa_est.add(inep)
        if dentro:
            arm_dentro.add(inep)
            if esfera == "Municipal": arm_dentro_mun.add(inep)
            elif esfera == "Estadual": arm_dentro_est.add(inep)
        if fora:
            arm_fora.add(inep)
            if esfera == "Municipal": arm_fora_mun.add(inep)
            elif esfera == "Estadual": arm_fora_est.add(inep)
        if sem_arm:
            nao_ha.add(inep)
            if esfera == "Municipal": nao_ha_mun.add(inep)
            elif esfera == "Estadual": nao_ha_est.add(inep)

    return {
        "despensa": despensa,
        "despensa_mun": despensa_mun,
        "despensa_est": despensa_est,

        "arm_dentro": arm_dentro,
        "arm_dentro_mun": arm_dentro_mun,
        "arm_dentro_est": arm_dentro_est,

        "arm_fora": arm_fora,
        "arm_fora_mun": arm_fora_mun,
        "arm_fora_est": arm_fora_est,

        "nao_ha": nao_ha,
        "nao_ha_mun": nao_ha_mun,
        "nao_ha_est": nao_ha_est,
    }

# ========================= CÁLCULO: IRREGULARIDADES DOS ALIMENTOS ARMAZENADOS =========================
def _calc_irr_armazenamento_alim(municipio: str, df: pd.DataFrame) -> Dict[dict, Set[str]]:
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_tag= _col_letter_to_index(COLS_AS["ALM_IRR"])

    fp, fp_mun, fp_est = set(), set(), set()
    sem_etq, sem_etq_mun, sem_etq_est = set(), set(), set()
    mofados, mofados_mun, mofados_est = set(), set(), set()
    exposto, exposto_mun, exposto_est = set(), set(), set()


    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        alimentos = str(row.iloc[i_tag])

        fora_prazo = "Alimentos fora do prazo de validade" in alimentos
        sem_id = "Ausência de etiquetas de identificação de validade" in alimentos
        podres = "Alimentos visivelmente mofados ou podres" in alimentos
        danificados = "Embalagens abertas, rasgadas ou danificadas expondo os alimentos à contaminação" in alimentos

        if fora_prazo:
            fp.add(inep)
            if esfera == "Municipal": fp_mun.add(inep)
            elif esfera == "Estadual": fp_est.add(inep)
        if sem_id:
            sem_etq.add(inep)
            if esfera == "Municipal": sem_etq.add(inep)
            elif esfera == "Estadual": sem_etq.add(inep)
        if podres:
            mofados.add(inep)
            if esfera == "Municipal": mofados_mun.add(inep)
            elif esfera == "Estadual": mofados_est.add(inep)
        if danificados:
            exposto.add(inep)
            if esfera == "Municipal": exposto_mun.add(inep)
            elif esfera == "Estadual": exposto_est.add(inep)
 
    return {
        "fora_prazo": fp,
        "fora_prazo_mun": fp_mun,
        "fora_prazo_est": fp_est,

        "sem_etq": sem_etq,
        "sem_etq_mun": sem_etq_mun,
        "sem_etq_est": sem_etq_est,

        "mofados": mofados,
        "mofados_mun": mofados_mun,
        "mofados_est": mofados_est,

        "exposto": exposto,
        "exposto_mun": exposto_mun,
        "exposto_est": exposto_est,
    }

# ========================= CÁLCULO: ULTRAPROCESSADOS =========================
def _ultraprocessados_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("ARMZ_UP", "").replace("<<", "").replace(">>", "")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    # Remove a palavra fraldas para evitar overlap com outras tags
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _ultraprocessados_match(er_norm: str, acesso_key: str) -> bool:
    toks = _ultraprocessados_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_ultraprocessados_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["ALM_UP"]) #Substituir pela coluna
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _ultraprocessados_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: CONSUMO DE ULTRAPROCESSADOS =========================
def _cons_ultraprocessados_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("UP_CONSUMO", "").replace("<<", "").replace(">>", "")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    # Remove a palavra fraldas para evitar overlap com outras tags
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _cons_ultraprocessados_match(er_norm: str, acesso_key: str) -> bool:
    toks = _cons_ultraprocessados_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_cons_ultraprocessados_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["CONS_UP"]) #Substituir pela coluna
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _cons_ultraprocessados_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: VENDA DE ULTRAPROCESSADOS =========================
def _venda_ultraprocessados_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("UP_COMERCIO", "").replace("<<", "").replace(">>", "")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    # Remove a palavra fraldas para evitar overlap com outras tags
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _venda_ultraprocessados_match(er_norm: str, acesso_key: str) -> bool:
    toks = _venda_ultraprocessados_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_venda_ultraprocessados_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["VEND_UP"]) #Substituir pela coluna
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _venda_ultraprocessados_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: DIRETRIZ ULTRAPROCESSADOS =========================
def _dir_ultraprocessados_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("UP_DIRETRIZ", "").replace("<<", "").replace(">>", "")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    # Remove a palavra fraldas para evitar overlap com outras tags
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _dir_ultraprocessados_match(er_norm: str, acesso_key: str) -> bool:
    toks = _dir_ultraprocessados_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_dir_ultraprocessados_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["DIR_UP"]) #Substituir pela coluna
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _dir_ultraprocessados_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: ARMAZENAMENTO DE ALIMENTOS CONGELADOS =========================
def _armz_congelados_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("ARMZ_CONGELADOS", "").replace("<<", "").replace(">>", "")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    # Remove a palavra fraldas para evitar overlap com outras tags
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _armz_congelados_match(er_norm: str, acesso_key: str) -> bool:
    toks = _armz_congelados_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_armz_congelados_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["ALM_CONG"]) #Substituir pela coluna
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _armz_congelados_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: IRREGULARIDADES NO ARMAZENAMENTO DE ALIMENTOS CONGELADOS =========================
def _irr_congelados_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("ARMZ_CONG_IRREG", "").replace("<<", "").replace(">>", "").replace("ENCONTRADO","")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    # Remove a palavra fraldas para evitar overlap com outras tags
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _irr_congelados_match(er_norm: str, acesso_key: str) -> bool:
    toks = _irr_congelados_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_irr_congelados_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["IRR_CONG"]) #Substituir pela coluna
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _irr_congelados_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: CARDÁPIO =========================
def _cardapio_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("CARDAPIO", "").replace("<<", "").replace(">>", "")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    # Remove a palavra fraldas para evitar overlap com outras tags
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _cardapio_match(er_norm: str, acesso_key: str) -> bool:
    toks = _cardapio_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_cardapio_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["CARD"]) #Substituir pela coluna
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        er_txt = er_txt.replace("Não","inexistente").replace("Sim, NÃO assinado por nutricionista e NÃO fixado em local visível","NAONAO").replace("Sim, NÃO assinado por nutricionista, mas fixado em local visível","naomas")
        
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _cardapio_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: CARDÁPIO ESPECIAL =========================
def _calc_cardapio_especial(municipio: str, df: pd.DataFrame) -> Dict[dict, Set[str]]:
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_tag= _col_letter_to_index(COLS_AS["CARD_ESP"])

    sim, sim_mun, sim_est = set(), set(), set()
    sim_irr, sim_irr_mun, sim_irr_est = set(), set(), set()
    nao, nao_mun, nao_est = set(), set(), set()

    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        cardapios = str(row.iloc[i_tag])

        tem_card = "Sim, havendo cardápio especial para esses alunos" in cardapios
        tem_card_irr = "Sim, mas não havendo cardápio especial para esses alunos" in cardapios
        nao_ha = "Não" in cardapios

        if tem_card:
            sim.add(inep)
            if esfera == "Municipal": sim_mun.add(inep)
            elif esfera == "Estadual": sim_est.add(inep)
        if tem_card_irr:
            sim_irr.add(inep)
            if esfera == "Municipal": sim_irr_mun.add(inep)
            elif esfera == "Estadual": sim_irr_est.add(inep)
        if nao_ha:
            nao.add(inep)
            if esfera == "Municipal": nao_mun.add(inep)
            elif esfera == "Estadual": nao_est.add(inep)
 
    return {
        "sim": sim,
        "sim_mun": sim_mun,
        "sim_est": sim_est,

        #tem criança com restrição alimentar MAS não possui cardápio especial
        "sim_irr": sim_irr,
        "sim_irr_mun": sim_irr_mun,
        "sim_irr_est": sim_irr_est,

        # Não há crianças com restrição alimentar
        "nao": nao,
        "nao_mun": nao_mun,
        "nao_est": nao_est,
    }

# ========================= CÁLCULO: REFEIÇÃO SERVIDA =========================
def _calc_refeicao_servida(municipio: str, df: pd.DataFrame) -> Dict[dict, Set[str]]:
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_tag= _col_letter_to_index(COLS_AS["REF_SERV"])

    #serv-servida; prep-preparada
    serv, serv_mun, serv_est = set(), set(), set()
    prep, prep_mun, prep_est = set(), set(), set()
    nao, nao_mun, nao_est = set(), set(), set()

    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        refeicoes = str(row.iloc[i_tag])

        servida = "Sim, havia refeição sendo SERVIDA" in refeicoes
        preparada = "Sim, havia refeição sendo PREPARADA" in refeicoes
        nao_ha = "Não" in refeicoes

        if servida:
            serv.add(inep)
            if esfera == "Municipal": serv_mun.add(inep)
            elif esfera == "Estadual": serv_est.add(inep)
        if preparada:
            prep.add(inep)
            if esfera == "Municipal": prep_mun.add(inep)
            elif esfera == "Estadual": prep_est.add(inep)
        if nao_ha:
            nao.add(inep)
            if esfera == "Municipal": nao_mun.add(inep)
            elif esfera == "Estadual": nao_est.add(inep)
 
    return {
        #Se tinha refeição sendo SERVIDA na hora da visita
        "serv": serv,
        "serv_mun": serv_mun,
        "serv_est": serv_est,

        #se tinha refeição sendo PREPARADA
        "prep": prep,
        "prep_mun": prep_mun,
        "prep_est": prep_est,

        "nao": nao,
        "nao_mun": nao_mun,
        "nao_est": nao_est,
    }

# ========================= CÁLCULO: CARDÁPIO DE ACORDO =========================
def _calc_cardapio_conforme(municipio: str, df: pd.DataFrame) -> Dict[dict, Set[str]]:
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_tag= _col_letter_to_index(COLS_AS["CARD_CONF"])

    #serv-servida; prep-preparada
    todos, todos_mun, todos_est = set(), set(), set()
    alguns, alguns_mun, alguns_est = set(), set(), set()
    nao, nao_mun, nao_est = set(), set(), set()

    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        cardapios = str(row.iloc[i_tag])

        todos_prev = "Sim, todos os itens previstos no cardápio" in cardapios
        alguns_prev = "Sim, mas apenas alguns itens previstos no cardápio" in cardapios
        nenhum = "Não, nenhum dos itens previstos no cardápio" in cardapios

        if todos_prev:
            todos.add(inep)
            if esfera == "Municipal": todos_mun.add(inep)
            elif esfera == "Estadual": todos_est.add(inep)
        if alguns_prev:
            alguns.add(inep)
            if esfera == "Municipal": alguns_mun.add(inep)
            elif esfera == "Estadual": alguns_est.add(inep)
        if nenhum:
            nao.add(inep)
            if esfera == "Municipal": nao_mun.add(inep)
            elif esfera == "Estadual": nao_est.add(inep)
 
    return {
        
        "conf_todos":"todos",
        "conf_todos_mun":"todos_mun",
        "conf_todos_est":"todos_est",

        "conf_alguns":"alguns",
        "conf_alguns_mun":"alguns_mun",
        "conf_alguns_est":"alguns_est",

        "nenhum":"nao",
        "nenhum_mun":"nao_mun",
        "nenhum_est":"nao_est",
    }

# ========================= CÁLCULO: IRREGULARIDADES NO REFETÓRIO =========================
def _irr_refeitorio_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("REFEITORIO_INFRA", "").replace("<<", "").replace(">>", "").replace("ILUM","ILUMINACAO").replace("VENT","VENTILACAO")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    # Remove a palavra fraldas para evitar overlap com outras tags
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _irr_refeitorio_match(er_norm: str, acesso_key: str) -> bool:
    toks = _irr_refeitorio_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_irr_refeitorio_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["IRR_REF"]) #Substituir pela coluna
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c").replace("nao foram identificados os aspectos irregulares listados acima","ok")
        
        # Verifica se a dependência está presente
        if _irr_refeitorio_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: IRREGULARIDADES DE ARMAZENAMENTO =========================
def _irr_armazenamento_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("ARMZ_LOCAL_IRREG", "").replace("<<", "").replace(">>", "")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _irr_armazenamento_match(er_norm: str, acesso_key: str) -> bool:
    toks = _irr_armazenamento_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_irr_armazenamento_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["ARM_IRR"])
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c").replace("","")
        
        # Verifica se a dependência está presente
        if _irr_armazenamento_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}


# ========================= CÁLCULO: ACESSIBILIDADE (RAMPAS, VÃO DE ENTRADA E PORTAS) =========================
def _calc_irregularidades_entrada(municipio: str, df: pd.DataFrame) -> Dict[dict, Set[str]]:
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_irregularidade = _col_letter_to_index(COLS_AS["IRR_ENTRADA"])

    sem_rampas, sem_rampas_mun, sem_rampas_est = set(), set(), set()
    rampas_irregulares, rampas_irregulares_mun, rampas_irregulares_est = set(), set(), set()
    vao_entrada, vao_entrada_mun, vao_entrada_est = set(), set(), set()

    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        irregularidades = str(row.iloc[i_irregularidade])

        nao_tem = "Não há rampa de acesso" in irregularidades
        rampa_irregular = "Há rampa de acesso, mas ela apresenta alguma irregularidade" in irregularidades
        vao_irregular = "Não há porta de entrada com largura de vão livre igual ou superior a 80cm" in irregularidades

        if nao_tem:
            sem_rampas.add(inep)
            if esfera == "Municipal": sem_rampas_mun.add(inep)
            elif esfera == "Estadual": sem_rampas_est.add(inep)
        if rampa_irregular:
            rampas_irregulares.add(inep)
            if esfera == "Municipal": rampas_irregulares_mun.add(inep)
            elif esfera == "Estadual": rampas_irregulares_est.add(inep)
        if vao_irregular:
            vao_entrada.add(inep)
            if esfera == "Municipal": vao_entrada_mun.add(inep)
            elif esfera == "Estadual": vao_entrada_est.add(inep)

    return {
        "sem_rampas": sem_rampas,
        "sem_rampas_mun": sem_rampas_mun,
        "sem_rampas_est": sem_rampas_est,

        "rampas_irregulares": rampas_irregulares,
        "rampas_irregulares_mun": rampas_irregulares_mun,
        "rampas_irregulares_est": rampas_irregulares_est,

        "sem_vao_entrada": vao_entrada,
        "sem_vao_entrada_mun": vao_entrada_mun,
        "sem_vao_entrada_est": vao_entrada_est,
    }

# ========================= CÁLCULO: VIGILÂNCIA SANITÁRIA =========================
def _calc_vigilancia_sanitaria(municipio: str, df: pd.DataFrame) -> Dict[str, Set[str]]:
    # TODO add AVCB e outros certifs aq
    """Calcula estatísticas de vigilância sanitária (ANVISA, AVCB, etc.)."""
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_anvisa = _col_letter_to_index(COLS_AS["ANVISA"])
    i_avcb = _col_letter_to_index(COLS_AS["AVCB"])
    i_dedet = _col_letter_to_index(COLS_AS["DEDET"])

    # v->válido; fv->fora de validade; nao->não tem 
    # ANVISA
    anvisa_valido, anvisa_valido_mun, anvisa_valido_est = set(), set(), set()
    anvisa_fv, anvisa_fv_mun, anvisa_fv_est = set(), set(), set()
    sem_anvisa, sem_anvisa_mun, sem_anvisa_est = set(), set(), set()

    # ANVISA
    avcb_valido, avcb_valido_mun, avcb_valido_est = set(), set(), set()
    avcb_fv, avcb_fv_mun, avcb_fv_est = set(), set(), set()
    sem_avcb, sem_avcb_mun, sem_avcb_est = set(), set(), set()

    # DEDETIZAÇÃO
    # dp-> dentro do prazo (6 meses) fp->fora do prazo (6 meses)
    dedet_dp, dedet_dp_mun, dedet_dp_est = set(), set(), set()
    dedet_fp, dedet_fp_mun, dedet_fp_est = set(), set(), set()
    sem_dedet, sem_dedet_mun, sem_dedet_est = set(), set(), set()


    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        
        anvisa = str(row.iloc[i_anvisa])
        avcb = str(row.iloc[i_avcb])
        dedetizacao = str(row.iloc[i_dedet])
        
        # <-- ANVISA -->
        valida = "Sim, válida" in anvisa
        nao_tem    = "Não" in anvisa
        fora_validade   = "Sim, mas está fora da validade"  in anvisa

        # <-- AVCb -->
        avcb_valida = "Sim, válida" in avcb
        avcb_fora_validade   = "Sim, mas está fora da validade"  in avcb
        nao_tem_avcb    = "Não" in avcb
        
        # <-- DEDETIZAÇÃO -->
        dentro_prazo = "Sim, emitido há no máximo 6 meses" in dedetizacao
        fora_prazo = "Sim, emitido há mais de 6 meses" in dedetizacao
        nao_dedet = "Não" in dedetizacao

        if valida:
            anvisa_valido.add(inep)
            if esfera == "Municipal":
                anvisa_valido_mun.add(inep)
            elif esfera == "Estadual":
                anvisa_valido_est.add(inep)
        if fora_validade:
            anvisa_fv.add(inep)
            if esfera == "Municipal": anvisa_fv_mun.add(inep)
            elif esfera == "Estadual": anvisa_fv_est.add(inep)
        if nao_tem:
            sem_anvisa.add(inep)
            if esfera == "Municipal": sem_anvisa_mun.add(inep)
            elif esfera == "Estadual": sem_anvisa_est.add(inep)

        if avcb_valida:
            avcb_valido.add(inep)
            if esfera == "Municipal": avcb_valido_mun.add(inep)
            elif esfera == "Estadual": avcb_valido_est.add(inep)
        if avcb_fora_validade:
            avcb_fv.add(inep)
            if esfera == "Municipal": avcb_fv_mun.add(inep)
            elif esfera == "Estadual": avcb_fv_est.add(inep)
        if nao_tem_avcb:
            sem_avcb.add(inep)
            if esfera == "Municipal": sem_avcb_mun.add(inep)
            elif esfera == "Estadual": sem_avcb_est.add(inep)

        if dentro_prazo:
            dedet_dp.add(inep)
            if esfera == "Municipal": dedet_dp_mun.add(inep)
            elif esfera == "Estadual": dedet_dp_est.add(inep)
        if fora_prazo:
            dedet_fp.add(inep)
            if esfera == "Municipal": dedet_fp_mun.add(inep)
            elif esfera == "Estadual": dedet_fp_est.add(inep)
        if nao_dedet:
            sem_dedet.add(inep)
            if esfera == "Municipal": sem_dedet_mun.add(inep)
            elif esfera == "Estadual": sem_dedet_est.add(inep)
        
        
    return {
        "com_anvisa": anvisa_valido,
        "com_anvisa_mun":anvisa_valido_mun,
        "com_anvisa_est": anvisa_valido_est,

        "fora_validade": anvisa_fv,
        "fora_validade_mun": anvisa_fv_mun,
        "fora_validade_est": anvisa_fv_est,

        "sem_anvisa": sem_anvisa,
        "sem_anvisa_mun": sem_anvisa_mun,
        "sem_anvisa_est": sem_anvisa_est,

        # === AVCb ===
        "com_avcb": avcb_valido,
        "com_avcb_mun":avcb_valido_mun,
        "com_avcb_est": avcb_valido_est,

        "avcb_fora_validade": avcb_fv,
        "avcb_fora_validade_mun": avcb_fv_mun,
        "avcb_fora_validade_est": avcb_fv_est,

        "sem_avcb": sem_avcb,
        "sem_avcb_mun": sem_avcb_mun,
        "sem_avcb_est": sem_avcb_est,

        # === DEDETIZAÇÃO ===
        "dedet_dp": dedet_dp,
        "dedet_dp_mun": dedet_dp_mun,
        "dedet_dp_est": dedet_dp_est,

        "dedet_fp": dedet_fp,
        "dedet_fp_mun": dedet_fp_mun,
        "dedet_fp_est": dedet_fp_est,

        "sem_dedet": sem_dedet,
        "sem_dedet_mun": sem_dedet_mun,
        "sem_dedet_est": sem_dedet_est,
    }
# ========================= ABASTECIMENTO DE ÁGUA =========================
def _calc_abastecimento_agua(municipio: str, df: pd.DataFrame) -> Dict[str, Set[str]]:
    # TODO add AVCB e outros certifs aq
    """Calcula estatísticas de vigilância sanitária (ANVISA, AVCB, etc.)."""
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_agua = _col_letter_to_index(COLS_AS["ABS_AGUA"])

    rede_todas, rede_mun, rede_est = set(), set(), set()
    poco_todas, poco_mun, poco_est = set(), set(), set()
    cacimba_todas, cacimba_mun, cacimba_est = set(), set(), set()
    fonte_todas, fonte_mun, fonte_est = set(), set(), set()
    sem_agua_todas, sem_agua_mun, sem_agua_est = set(), set(), set()

    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        
        abastecimento = str(row.iloc[i_agua])
        
        rede = "Rede Pública" in abastecimento
        poco    = "Poço artesiano" in abastecimento
        cacimba = "Cacimba/cisterna/poço" in abastecimento
        fonte   = "Fonte/rio/igarapé/córrego" in abastecimento
        nao_ha = "Não há" in abastecimento

        if rede:
            rede_todas.add(inep)
            if esfera == "Municipal": rede_mun.add(inep)
            elif esfera == "Estadual": rede_est.add(inep)
        if poco:
            poco_todas.add(inep)
            if esfera == "Municipal": poco_mun.add(inep)
            elif esfera == "Estadual": poco_est.add(inep)
        if cacimba:
            cacimba_todas.add(inep)
            if esfera == "Municipal": cacimba_mun.add(inep)
            elif esfera == "Estadual": cacimba_est.add(inep)
        if fonte:
            fonte_todas.add(inep)
            if esfera == "Municipal": fonte_mun.add(inep)
            elif esfera == "Estadual": fonte_est.add(inep)  
        if nao_ha:
            sem_agua_todas.add(inep)
            if esfera == "Municipal": sem_agua_mun.add(inep)
            elif esfera == "Estadual": sem_agua_est.add(inep)
        
        
        
    return {
        "rede_publica": rede_todas,
        "rede_publica_mun": rede_mun,
        "rede_publica_est": rede_est,
 

        "poco_artesiano": poco_todas,
        "poco_artesiano_mun": poco_mun,
        "poco_artesiano_est": poco_est,

        "cacimba_etc": cacimba_todas,
        "cacimba_etc_mun": cacimba_mun,
        "cacimba_etc_est": cacimba_est,
        

        "fonte_etc": fonte_todas,
        "fonte_etc_mun": fonte_mun,
        "fonte_etc_est": fonte_est,

        "sem_agua": sem_agua_todas,
        "sem_agua_mun": sem_agua_mun,
        "sem_agua_est": sem_agua_est,
    }

# ========================= RESERVATÓRIO DE ÁGUA =========================

def _reservatorio_agua_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("RESERVATORIO", "").replace("<<", "").replace(">>", "").replace("FUNC","FUNCIONAMENTO").replace("CISTERNA_OK","CISREGULAR")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    # Remove a palavra fraldas para evitar overlap com outras tags
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _reservatorio_agua_match(er_norm: str, acesso_key: str) -> bool:
    toks = _reservatorio_agua_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_reservatorio_agua_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["RES_AGUA"]) #Substituir pela coluna
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c").replace(" em funcionamento","ok").replace("cisterna,ok","cisregular")
        
        # Verifica se a dependência está presente
        if _reservatorio_agua_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= SISTEMA DE ESGOTAMENTO =========================

def _calc_sistema_esgotamento(municipio: str, df: pd.DataFrame) -> Dict[str, Set[str]]:
    # TODO add AVCB e outros certifs aq
    """Calcula estatísticas de vigilância sanitária (ANVISA, AVCB, etc.)."""
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_esg = _col_letter_to_index(COLS_AS["ESGOTAMENTO"])

    esgot_san, esgot_san_mun, esgot_san_est = set(), set(), set()
    fossa_todas, fossa_mun, fossa_est = set(), set(), set()
    despejo_inadeq, despejo_inadeq_mun, despejo_inadeq_est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        
        esgotamento = str(row.iloc[i_esg])
        
        conectado = "Conexão com rede de esgotamento sanitário" in esgotamento
        fossa     = "Fossa, sumidouro ou similar" in esgotamento
        despejo_inadequado = "Despejo sem destinação adequada" in esgotamento

        if conectado:
            esgot_san.add(inep)
            if esfera == "Municipal": esgot_san_mun.add(inep)
            elif esfera == "Estadual": esgot_san_est.add(inep)
        if fossa:
            fossa_todas.add(inep)
            if esfera == "Municipal": fossa_mun.add(inep)
            elif esfera == "Estadual": fossa_est.add(inep)
        if despejo_inadequado:
            despejo_inadeq.add(inep)
            if esfera == "Municipal": despejo_inadeq_mun.add(inep)
            elif esfera == "Estadual": despejo_inadeq_est.add(inep)
        
    return {
        "sist_conectado": esgot_san,
        "sist_conectado_mun": esgot_san_mun,
        "sist_conectado_est": esgot_san_est,

        "fossa_e_outros": fossa_todas,
        "fossa_e_outros_mun": fossa_mun,
        "fossa_e_outros_est": fossa_est,

        "despejo_inadequado": despejo_inadeq,
        "despejo_inadequado_mun": despejo_inadeq_mun,
        "despejo_inadequado_est": despejo_inadeq_est,
    }

# ========================= DEPEDÊNCIAS (TAGS DINÂMICAS) =========================
isCobert = False
def _dep_tokens_from_tag(tag_suffix: str) -> List[str]:
    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("DEP_", "").replace("<<", "").replace(">>", "").replace("PATIO_COBERTO","COBERT").replace("PATIO_DESCOBERTO","DESCOB")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _dep_match(er_norm: str, dep_key: str) -> bool:
    toks = _dep_tokens_from_tag(dep_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_dependencias_por_tag(municipio: str, df: pd.DataFrame, dep_key: str) -> Dict[str, Set[str]]:
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_er = _col_letter_to_index(COLS_AS["DEPENDENCIAS"])
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_er])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c").replace("patio coberto"," cobert").replace("patio descoberto"," descob")
        
        # Verifica se a dependência está presente
        if _dep_match(er_norm, dep_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: DESTINAÇÃO DO LIXO =========================
def _destinacao_lixo_tokens_from_tag(tag_suffix: str) -> List[str]:
    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("LIXO_", "").replace("<<", "").replace(">>", "")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _destinacao_lixo_match(er_norm: str, acesso_key: str) -> bool:
    toks = _destinacao_lixo_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_destinacao_lixo_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["LIXO"])
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _destinacao_lixo_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: CONEXÃO REDE ELÉTRICA =========================
def _calc_energia(municipio: str, df: pd.DataFrame) -> Dict[dict, Set[str]]:
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_tag= _col_letter_to_index(COLS_AS["ENERGIA"])

    sim_func, sim_func_mun, sim_func_est = set(), set(), set()
    sim_fora_func, sim_fora_func_mun, sim_fora_func_est = set(), set(), set()
    nao, nao_mun, nao_est = set(), set(), set()

    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        conexao = str(row.iloc[i_tag])

        funciona = "Sim, em funcionamento" in conexao
        fora_funciona = "Sim, mas fora de funcionamento" in conexao
        nao_ha = "Não" in conexao

        if funciona:
            sim_func.add(inep)
            if esfera == "Municipal": sim_func_mun.add(inep)
            elif esfera == "Estadual": sim_func_est.add(inep)
        if fora_funciona:
            sim_fora_func.add(inep)
            if esfera == "Municipal": sim_fora_func_mun.add(inep)
            elif esfera == "Estadual": sim_fora_func_est.add(inep)
        if nao_ha:
            nao.add(inep)
            if esfera == "Municipal": nao_mun.add(inep)
            elif esfera == "Estadual": nao_est.add(inep)
 
    return {
        "em_funcionamento": sim_func,
        "em_funcionamento_mun": sim_func_mun,
        "em_funcionamento_est": sim_func_est,

        "fora_funcionamento": sim_fora_func,
        "fora_funcionamento_mun": sim_fora_func_mun,
        "fora_funcionamento_est": sim_fora_func_est,

        "nao": nao,
        "nao_mun": nao_mun,
        "nao_est": nao_est,
    }

# ========================= CÁLCULO: SALAS IRREGULARES =========================
# TODO mudar tratamento para sala irregular
def _irregularidades_salas_tokens_from_tag(tag_suffix: str) -> List[str]:
    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("SALAS_IRREGULARES", "").replace("<<", "").replace(">>", "")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA", "AUSÊNCIA", "OU"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _irregularidades_salas_match(er_norm: str, dep_key: str) -> bool:
    toks = _irregularidades_salas_tokens_from_tag(dep_key)
    
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_irregularidades_salas_por_tag(municipio: str, df: pd.DataFrame, dep_key: str) -> Dict[str, Set[str]]:
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_er = _col_letter_to_index(COLS_AS["SALA_IRR"]) 
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_er])
        
        if not er_txt:
            continue
        
        # pula questões que não tem irregularidades
        if(er_txt == 'Não foram identificados os aspectos irregulares listados acima'):
            continue

        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _irregularidades_salas_match(er_norm, dep_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: BRINQUEDOS DOS PARQUINHOS =========================
def _calc_brinquedos_parquinho(municipio: str, df: pd.DataFrame) -> Dict[dict, Set[str]]:
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_irregularidade = _col_letter_to_index(COLS_AS["PARQ_INF"])

    sem_parq, sem_parq_mun, sem_parq_est = set(), set(), set()
    cond, cond_mun, cond_est = set(), set(), set()
    sem_cond, sem_cond_mun, sem_cond_est = set(), set(), set()

    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        irregularidades = str(row.iloc[i_irregularidade])

        nao_tem = "Não" in irregularidades
        com_cond = "Sim, em condições de uso" in irregularidades
        nao_tem_cond = "Sim, mas sem condições de uso" in irregularidades

        if nao_tem:
            sem_parq.add(inep)
            if esfera == "Municipal": sem_parq_mun.add(inep)
            elif esfera == "Estadual": sem_parq_est.add(inep)
        if com_cond:
            cond.add(inep)
            if esfera == "Municipal": cond_mun.add(inep)
            elif esfera == "Estadual": cond_est.add(inep)
        if nao_tem_cond:
            sem_cond.add(inep)
            if esfera == "Municipal": sem_cond_mun.add(inep)
            elif esfera == "Estadual": sem_cond_est.add(inep)
 
 

    return {
        "nao": sem_parq,
        "nao_mun": sem_parq_mun,
        "nao_est": sem_parq_est,

        "cond": cond,
        "cond_mun": cond_mun,
        "cond_est": cond_est,

        "sem_cond": sem_cond,
        "sem_cond_mun": sem_cond_mun,
        "sem_cond_est": sem_cond_est,
    }

# ========================= CÁLCULO: IRREGULARIDADES NA BIBLIOTECAS =========================
#substituir i pelo nome da tag calculada
def _irr_biblioteca_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("BIB_EXC_INFRA", "").replace("<<", "").replace(">>", "").replace("AUS","AUSENCIA").replace("NENHUMA","ASPECTOS").replace("IRREGULARES","").replace("IRREGULAR","")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    # Remove a palavra fraldas para evitar overlap com outras tags
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _irr_biblioteca_match(er_norm: str, acesso_key: str) -> bool:
    toks = _irr_biblioteca_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_irr_biblioteca_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["IRR_BIB"]) #Substituir pela coluna
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _irr_biblioteca_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: IRREGULARIDADES NA SALA DE LEITURA =========================
#substituir i pelo nome da tag calculada
def _irr_sala_leitura_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("SL_EXC_INFRA", "").replace("<<", "").replace(">>", "").replace("AUS","AUSENCIA").replace("NENHUMA","ASPECTOS").replace("IRREGULARES","").replace("IRREGULAR","")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    # Remove a palavra fraldas para evitar overlap com outras tags
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _irr_sala_leitura_match(er_norm: str, acesso_key: str) -> bool:
    toks = _irr_sala_leitura_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_irr_sala_leitura_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["IRR_SL"]) #Substituir pela coluna
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _irr_sala_leitura_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: IRREGULARIDADES NA SALA DE LEITURA + BIBLIOTECA (ESPAÇO COMPARTILHADO) =========================
#substituir i pelo nome da tag calculada
def _irr_bib_compart_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("BIBLI_SL_COMP_INFRA", "").replace("<<", "").replace(">>", "").replace("AUS","AUSENCIA").replace("NENHUMA","ASPECTOS").replace("IRREGULARES","").replace("IRREGULAR","")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    # Remove a palavra fraldas para evitar overlap com outras tags
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _irr_bib_compart_match(er_norm: str, acesso_key: str) -> bool:
    toks = _irr_bib_compart_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_irr_bib_compart_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["IRR_BIBSL"]) #Substituir pela coluna
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _irr_bib_compart_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: AGUA NO SANITÁRIO EXCLUSIVO PARA EDUC INFANTIL =========================
def _agua_san_infantil_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("EI_SAN_EXC", "").replace("<<", "").replace(">>", "").replace("EXISTE","SIM")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    # Remove a palavra fraldas para evitar overlap com outras tags
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _agua_san_infantil_match(er_norm: str, acesso_key: str) -> bool:
    toks = _agua_san_infantil_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_agua_san_infantil_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["AG_SAN_EI"]) #Substituir pela coluna
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _agua_san_infantil_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: SALAS MULTISSERIADAS =========================
def _calc_salas_multisseriadas(municipio: str, df: pd.DataFrame) -> Dict[dict, Set[str]]:
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_tag= _col_letter_to_index(COLS_AS["MULTS"])

    sim, sim_mun, sim_est = set(), set(), set()

    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        salas = str(row.iloc[i_tag])

        multisseriado = "Sim" in salas

        if multisseriado:
            sim.add(inep)
            if esfera == "Municipal": sim_mun.add(inep)
            elif esfera == "Estadual": sim_est.add(inep)
 
    return {
        "sim": sim,
        "sim_mun": sim_mun,
        "sim_est": sim_est,
    }


# ========================= CÁLCULO: IRREGULARIDADES NO SANITÁRIO EXCLUSIVO PARA EDUC INFANTIL =========================
def _irr_san_infantil_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("EI_SAN_EXC_INFRA", "").replace("<<", "").replace(">>", "").replace("NENHUMA","NAO").replace("PROBLEMA","")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    # Remove a palavra fraldas para evitar overlap com outras tags
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _irr_san_infantil_match(er_norm: str, acesso_key: str) -> bool:
    toks = _irr_san_infantil_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_irr_san_infantil_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["IRR_SAN_EI"]) #Substituir pela coluna
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _irr_san_infantil_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: ITEM LACTÁRIO =========================
def _item_lactario_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("LACTARIO_ITEM", "").replace("<<", "").replace(">>", "")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _item_lactario_match(er_norm: str, acesso_key: str) -> bool:
    toks = _item_lactario_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_item_lactario_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["ITEM_LACT"])
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _item_lactario_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: IRREGULARIDADES LACTÁRIO =========================
def _irregularidades_lactario_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("LACTARIO_INFRA", "").replace("<<", "").replace(">>", "").replace("AUS","")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA","AUSÊNCIA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _irregularidades_lactario_match(er_norm: str, acesso_key: str) -> bool:
    toks = _irregularidades_lactario_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_irregularidades_lactario_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["IRR_LACT"])
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _irregularidades_lactario_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}
# ========================= CÁLCULO: LOCALIDADE FRALDÁRIO =========================
def _calc_localidade_fraldario(municipio: str, df: pd.DataFrame) -> Dict[dict, Set[str]]:
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_tag= _col_letter_to_index(COLS_AS["LOCAL_FRAL"])

    dentro, dentro_mun, dentro_est = set(), set(), set()
    separado, separado_mun, separado_est = set(), set(), set()

    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        localidade = str(row.iloc[i_tag])

        dentro_berc = "Implantado dentro do berçário" in localidade
        amb_separado = "Em ambiente separado" in localidade

        if dentro_berc:
            dentro.add(inep)
            if esfera == "Municipal": dentro_mun.add(inep)
            elif esfera == "Estadual": dentro_est.add(inep)
        if amb_separado:
            separado.add(inep)
            if esfera == "Municipal": separado_mun.add(inep)
            elif esfera == "Estadual": separado_est.add(inep)
 
    return {
        "local_dentro": dentro,
        "local_dentro_mun": dentro_mun,
        "local_dentro_est": dentro_est,

        "local_separado": separado,
        "local_separado_mun": separado_mun,
        "local_separado_est": separado_est,
    }

# ========================= CÁLCULO: ITEM FRALDÁRIO =========================
def _item_fraldario_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    
    sufixo = tag_suffix.replace("FRALDARIO_ITEM", "").replace("<<", "").replace(">>", "")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    # Remove a palavra fraldas para evitar overlap com outras tags
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA","FRALDAS"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _item_fraldario_match(er_norm: str, acesso_key: str) -> bool:
    toks = _item_fraldario_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_item_fraldario_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["ITEM_FRAL"])
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _item_fraldario_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: IRREGULARIDADES FRALDÁRIO =========================
def _irregularidades_fraldario_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("FRALDARIO_INFRA", "").replace("<<", "").replace(">>", "")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    # Remove a palavra fraldas para evitar overlap com outras tags
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _irregularidades_fraldario_match(er_norm: str, acesso_key: str) -> bool:
    toks = _irregularidades_fraldario_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_irregularidades_fraldario_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["IRR_FRAL"])
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _irregularidades_fraldario_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {"total": total, "mun": mun, "est": est}

# ========================= CÁLCULO: IRREGULARIDADES BERÇÁRIO =========================
def _irregularidades_bercario_tokens_from_tag(tag_suffix: str) -> List[str]:

    # Remove o prefixo comum se existir
    sufixo = tag_suffix.replace("BERCARIO_INFRA", "").replace("<<", "").replace(">>", "")
    
    # Substitui underscores por espaços
    sufixo = sufixo.replace("_", " ")
    
    # Quebra em tokens e normaliza
    toks = [t.strip() for t in re.split(r"\s+", sufixo) if t.strip()]
    
    # Remove tokens muito genéricos (palavras conectivas)
    # Mantemos palavras importantes como "AREA" por segurança
    # Remove a palavra fraldas para evitar overlap com outras tags
    stopwords = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "COM", "PARA","AUSÊNCIA"}
    toks = [t for t in toks if t.upper() not in stopwords]
    
    return [t for t in toks if t]

def _irregularidades_bercario_match(er_norm: str, acesso_key: str) -> bool:
    toks = _irregularidades_bercario_tokens_from_tag(acesso_key)
    
    # Testa inclusão de todos os tokens (como palavras soltas, ordem livre)
    # Normaliza para comparação sem acentos e case insensitive
    return all(
        t.lower().replace("ç", "c") in er_norm 
        for t in [re.sub(r"\s+", " ", _norm_noacc(x)) for x in toks]
    )

def _calc_irregularidades_bercario_por_tag(municipio: str, df: pd.DataFrame, acesso_key: str) -> Dict[str, Set[str]]:

    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_acesso = _col_letter_to_index(COLS_AS["IRR_BER"])
    
    total, mun, est = set(), set(), set()
    
    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        er_txt = str(row.iloc[i_acesso])
        
        if not er_txt:
            continue
        
        # Normaliza o texto de dependências (remove acentos, lowercase)
        er_norm = _norm_noacc(er_txt).lower().replace("ç", "c")
        
        # Verifica se a dependência está presente
        if _irregularidades_bercario_match(er_norm, acesso_key):
            total.add(inep)
            
            if esfera == "Municipal":
                mun.add(inep)
            elif esfera == "Estadual":
                est.add(inep)
    
    return {
        "total": total, 
        "mun": mun, 
        "est": est,
    }

# ========================= CÁLCULO: LIMITE BERÇÁRIO =========================
def _calc_limite_bercario(municipio: str, df: pd.DataFrame) -> Dict[dict, Set[str]]:
    i_mun = _col_letter_to_index(COLS_AS["MUNICIPIO"])
    i_inep = _col_letter_to_index(COLS_AS["INEP"])
    i_esf = _col_letter_to_index(COLS_AS["ESFERA"])
    i_tag= _col_letter_to_index(COLS_AS["LIM_BERC"])

    nao, nao_mun, nao_est = set(), set(), set()

    for _, row in df.iterrows():
        if str(row.iloc[i_mun]) != municipio:
            continue
        
        inep = str(row.iloc[i_inep]).strip()
        if not inep:
            continue
        
        esfera = str(row.iloc[i_esf]).strip()
        alimentos = str(row.iloc[i_tag])

        sem_etq = "Não" in alimentos

        if sem_etq:
            nao.add(inep)
            if esfera == "Municipal": nao_mun.add(inep)
            elif esfera == "Estadual": nao_est.add(inep)
 
    return {
        "nao": nao,
        "nao_mun": nao_mun,
        "nao_est": nao_est,
    }


# ========================= ********************* =========================
# ========================= PROCESSAMENTO DE TAGS =========================
# ========================= ********************* =========================
def calcular_tags_para_municipio(municipio: str, df: pd.DataFrame, tags: List[str]) -> Dict[str, str]:
    """Calcula todas as tags para um município específico."""
    print(f"  Processando município: {municipio}")
    resultados = {}
    muni_cache = {}  # Cache local para este município
    
    for tag in tags:
        # --------- ESCOLAS VISITADAS ----------
        if tag in {
            "<<NUMERO_ESCOLAS_VISITADAS>>",
            "<<NUMERO_ESCOLAS_VISITADAS_MUN>>",
            "<<NUMERO_ESCOLAS_VISITADAS_EST>>",
            "<<ESCOLAS_INFANTIL_VISITADAS>>",
            "<<ESCOLAS_INFANTIL_VISITADAS_MUN>>",
            "<<ESCOLAS_INFANTIL_VISITADAS_EST>>",
            "<<ESCOLAS_INFANTIL_VISITADAS_CRECHE>>",
            "<<ESCOLAS_INFANTIL_VISITADAS_CRECHE_MUN>>",
            "<<ESCOLAS_INFANTIL_VISITADAS_CRECHE_EST>>",
            "<<ESCOLAS_INFANTIL_VISITADAS_PREESCOLA>>",
            "<<ESCOLAS_INFANTIL_VISITADAS_PREESCOLA_MUN>>",
            "<<ESCOLAS_INFANTIL_VISITADAS_PREESCOLA_EST>>",
            "<<ESCOLAS_FUND_VISITADAS>>",
            "<<ESCOLAS_FUND_VISITADAS_MUN>>",
            "<<ESCOLAS_FUND_VISITADAS_EST>>",
            "<<ESCOLAS_INF_FUND_VISITADAS>>",
            "<<ESCOLAS_INF_FUND_VISITADAS_MUN>>",
            "<<ESCOLAS_INF_FUND_VISITADAS_EST>>",
            "<<ESCOLAS_MED_VISITADAS>>",
            "<<ESCOLAS_MED_VISITADAS_MUN>>",
            "<<ESCOLAS_MED_VISITADAS_EST>>",
            "<<ESCOLAS_EJA_VISITADAS>>",
            "<<ESCOLAS_EJA_VISITADAS_MUN>>",
            "<<ESCOLAS_EJA_VISITADAS_EST>>",
        }:
            if "esc_vis" not in muni_cache:
                muni_cache["esc_vis"] = _calc_escolas_visitadas(municipio, df)
            
            s = muni_cache["esc_vis"]
            def n(k): return str(len(s[k]))
            
            MAP = {
                "<<NUMERO_ESCOLAS_VISITADAS>>":      "todas",
                "<<NUMERO_ESCOLAS_VISITADAS_MUN>>":  "municipais",
                "<<NUMERO_ESCOLAS_VISITADAS_EST>>":  "estaduais",

                "<<ESCOLAS_INFANTIL_VISITADAS>>":        "infantil_q",
                "<<ESCOLAS_INFANTIL_VISITADAS_MUN>>":    "infantil_q_mun",
                "<<ESCOLAS_INFANTIL_VISITADAS_EST>>":    "infantil_q_est",

                "<<ESCOLAS_INFANTIL_VISITADAS_CRECHE>>":         "infantil_creche",
                "<<ESCOLAS_INFANTIL_VISITADAS_CRECHE_MUN>>":     "inf_creche_mun",
                "<<ESCOLAS_INFANTIL_VISITADAS_CRECHE_EST>>":     "inf_creche_est",

                "<<ESCOLAS_INFANTIL_VISITADAS_PREESCOLA>>":      "infantil_pre",
                "<<ESCOLAS_INFANTIL_VISITADAS_PREESCOLA_MUN>>":  "inf_pre_mun",
                "<<ESCOLAS_INFANTIL_VISITADAS_PREESCOLA_EST>>":  "inf_pre_est",

                "<<ESCOLAS_FUND_VISITADAS>>":         "fund_q",
                "<<ESCOLAS_FUND_VISITADAS_MUN>>":     "fund_q_mun",
                "<<ESCOLAS_FUND_VISITADAS_EST>>":     "fund_q_est",

                "<<ESCOLAS_INF_FUND_VISITADAS>>":         "inf_fund_exclusivo",
                "<<ESCOLAS_INF_FUND_VISITADAS_MUN>>":     "inf_fund_exclusivo_mun",
                "<<ESCOLAS_INF_FUND_VISITADAS_EST>>":     "inf_fund_exclusivo_est",

                "<<ESCOLAS_MED_VISITADAS>>":          "medio_q",
                "<<ESCOLAS_MED_VISITADAS_MUN>>":      "medio_q_mun",
                "<<ESCOLAS_MED_VISITADAS_EST>>":      "medio_q_est",

                "<<ESCOLAS_EJA_VISITADAS>>":          "eja_q",
                "<<ESCOLAS_EJA_VISITADAS_MUN>>":      "eja_q_mun",
                "<<ESCOLAS_EJA_VISITADAS_EST>>":      "eja_q_est",
            }
            
            key = MAP.get(tag)
            resultados[tag] = n(key) if key else ""
            continue


        # --------- IRREGULARIDADES NAS ENTRADAS ----------
        if tag in {
            "<<ESCOLAS_SEM_RAMPAS>>",
            "<<ESCOLAS_SEM_RAMPAS_MUN>>",
            "<<ESCOLAS_SEM_RAMPAS_EST>>",
            "<<ESCOLAS_RAMPAS_IRREGULARES>>",
            "<<ESCOLAS_RAMPAS_IRREGULARES_MUN>>",
            "<<ESCOLAS_RAMPAS_IRREGULARES_EST>>",
            "<<ESCOLAS_VAO_ENTRADA_IRREGULAR>>",
            "<<ESCOLAS_VAO_ENTRADA_IRREGULAR_MUN>>",
            "<<ESCOLAS_VAO_ENTRADA_IRREGULAR_EST>>",
        }:
            if "ent_fis" not in muni_cache:
                muni_cache["ent_fis"] = _calc_irregularidades_entrada(municipio, df)
            
            s = muni_cache["ent_fis"]
            def n(k): return str(len(s[k]))
            
            MAP = {
                "<<ESCOLAS_SEM_RAMPAS>>": "sem_rampas",
                "<<ESCOLAS_SEM_RAMPAS_MUN>>" : "sem_rampas_mun",
                "<<ESCOLAS_SEM_RAMPAS_EST>>": "sem_rampas_est",

                "<<ESCOLAS_RAMPAS_IRREGULARES>>": "rampas_irregulares",
                "<<ESCOLAS_RAMPAS_IRREGULARES_MUN>>": "rampas_irregulares_mun",
                "<<ESCOLAS_RAMPAS_IRREGULARES_EST>>": "rampas_irregulares_est",

                "<<ESCOLAS_VAO_ENTRADA_IRREGULAR>>": "sem_vao_entrada",
                "<<ESCOLAS_VAO_ENTRADA_IRREGULAR_MUN>>": "sem_vao_entrada_mun",
                "<<ESCOLAS_VAO_ENTRADA_IRREGULAR_EST>>": "sem_vao_entrada_est"
            }
            
            key = MAP.get(tag)
            resultados[tag] = n(key) if key else ""
            continue

        # --------- IRREGULARIDADES NOS ALIMENTOS ARMAZENADOS ----------
        if tag in {
            "<<ARMZ_ALIM_FORA_VALIDADE>>",
            "<<ARMZ_ALIM_FORA_VALIDADE_MUN>>",
            "<<ARMZ_ALIM_FORA_VALIDADE_EST>>",
            "<<ARMZ_ALIM_SEM_ETIQUETA_VAL>>",
            "<<ARMZ_ALIM_SEM_ETIQUETA_VAL_MUN>>",
            "<<ARMZ_ALIM_SEM_ETIQUETA_VAL_EST>>",
            "<<ARMZ_ALIM_MOFADOS_PODRES>>",
            "<<ARMZ_ALIM_MOFADOS_PODRES_MUN>>",
            "<<ARMZ_ALIM_MOFADOS_PODRES_EST>>",
            "<<ARMZ_ALIM_EMBALAGENS_DANIFICADAS>>",
            "<<ARMZ_ALIM_EMBALAGENS_DANIFICADAS_MUN>>",
            "<<ARMZ_ALIM_EMBALAGENS_DANIFICADAS_EST>>",
        }:
            if "irr_alim" not in muni_cache:
                muni_cache["irr_alim"] = _calc_irr_armazenamento_alim(municipio, df)
            
            s = muni_cache["irr_alim"]
            def n(k): return str(len(s[k]))
            
            MAP = {
                "<<ARMZ_ALIM_FORA_VALIDADE>>":"fora_prazo",
                "<<ARMZ_ALIM_FORA_VALIDADE_MUN>>":"fora_prazo_mun",
                "<<ARMZ_ALIM_FORA_VALIDADE_EST>>":"fora_prazo_est",

                "<<ARMZ_ALIM_SEM_ETIQUETA_VAL>>":"sem_etq",
                "<<ARMZ_ALIM_SEM_ETIQUETA_VAL_MUN>>":"sem_etq_mun",
                "<<ARMZ_ALIM_SEM_ETIQUETA_VAL_EST>>":"sem_etq_est",

                "<<ARMZ_ALIM_MOFADOS_PODRES>>":"mofados",
                "<<ARMZ_ALIM_MOFADOS_PODRES_MUN>>":"mofados_mun",
                "<<ARMZ_ALIM_MOFADOS_PODRES_EST>>":"mofados_est",

                "<<ARMZ_ALIM_EMBALAGENS_DANIFICADAS>>":"exposto",
                "<<ARMZ_ALIM_EMBALAGENS_DANIFICADAS_MUN>>":"exposto_mun",
                "<<ARMZ_ALIM_EMBALAGENS_DANIFICADAS_EST>>":"exposto_est",
            }
            
            key = MAP.get(tag)
            resultados[tag] = n(key) if key else ""
            continue
        
        # --------- VIGILÂNCIA SANITÁRIA ----------
        if tag in {
            "<<LIC_ANVISA_VALIDO>>",
            "<<LIC_ANVISA_VALIDO_MUN>>",
            "<<LIC_ANVISA_VALIDO_EST>>",
            "<<LIC_ANVISA_FORA_VAL>>",
            "<<LIC_ANVISA_FORA_VAL_MUN>>",
            "<<LIC_ANVISA_FORA_VAL_EST>>",
            "<<LIC_ANVISA_NAO>>",
            "<<LIC_ANVISA_NAO_MUN>>",
            "<<LIC_ANVISA_NAO_EST>>",
            "<<AVCB_VALIDO>>",
            "<<AVCB_VALIDO_MUN>>",
            "<<AVCB_VALIDO_EST>>",
            "<<AVCB_FORA_VAL>>",
            "<<AVCB_FORA_VAL_MUN>>",
            "<<AVCB_FORA_VAL_EST>>",
            "<<AVCB_NAO>>",
            "<<AVCB_NAO_MUN>>",
            "<<AVCB_NAO_EST>>",
            "<<DEDET_ATE6M>>",
            "<<DEDET_ATE6M_MUN>>",
            "<<DEDET_ATE6M_EST>>",
            "<<DEDET_MAIS6M>>",
            "<<DEDET_MAIS6M_MUN>>",
            "<<DEDET_MAIS6M_EST>>",
            "<<DEDET_NAO>>",
            "<<DEDET_NAO_MUN>>",
            "<<DEDET_NAO_EST>>",
        }:
            if "vig_san" not in muni_cache:
                muni_cache["vig_san"] = _calc_vigilancia_sanitaria(municipio, df)
            
            s = muni_cache["vig_san"]
            def n(k): return str(len(s[k]))
            
            MAP = {
                "<<LIC_ANVISA_VALIDO>>":      "com_anvisa",
                "<<LIC_ANVISA_VALIDO_MUN>>":  "com_anvisa_mun",
                "<<LIC_ANVISA_VALIDO_EST>>":  "com_anvisa_est",
                "<<LIC_ANVISA_FORA_VAL>>":      "fora_validade",
                "<<LIC_ANVISA_FORA_VAL_MUN>>":  "fora_validade_mun",
                "<<LIC_ANVISA_FORA_VAL_EST>>":  "fora_validade_est",
                "<<LIC_ANVISA_NAO>>":      "sem_anvisa",
                "<<LIC_ANVISA_NAO_MUN>>":  "sem_anvisa_mun",
                "<<LIC_ANVISA_NAO_EST>>":  "sem_anvisa_est",

                "<<AVCB_VALIDO>>":"com_avcb",
                "<<AVCB_VALIDO_MUN>>":"com_avcb_mun",
                "<<AVCB_VALIDO_EST>>":"com_avcb_est",
                "<<AVCB_FORA_VAL>>":"avcb_fora_validade",
                "<<AVCB_FORA_VAL_MUN>>":"avcb_fora_validade_mun",
                "<<AVCB_FORA_VAL_EST>>":"avcb_fora_validade_est",
                "<<AVCB_NAO>>":"sem_avcb",
                "<<AVCB_NAO_MUN>>":"sem_avcb_mun",
                "<<AVCB_NAO_EST>>":"sem_avcb_est",

                "<<DEDET_ATE6M>>":"dedet_dp",
                "<<DEDET_ATE6M_MUN>>":"dedet_dp_mun",
                "<<DEDET_ATE6M_EST>>":"dedet_dp_est",
                "<<DEDET_MAIS6M>>":"dedet_fp",
                "<<DEDET_MAIS6M_MUN>>":"dedet_fp_mun",
                "<<DEDET_MAIS6M_EST>>":"dedet_fp_est",
                "<<DEDET_NAO>>":"sem_dedet",
                "<<DEDET_NAO_MUN>>":"sem_dedet_mun",
                "<<DEDET_NAO_EST>>":"sem_dedet_est",
            }
            
            key = MAP.get(tag)
            resultados[tag] = n(key) if key else ""
            continue

        # --------- ABASTECIMENTO DE AGUA ----------
        if tag in {
            "<<AGUA_REDE_PUBLICA>>",
            "<<AGUA_REDE_PUBLICA_MUN>>",
            "<<AGUA_REDE_PUBLICA_EST>>",
            "<<AGUA_POCO_ARTESIANO>>",
            "<<AGUA_POCO_ARTESIANO_MUN>>",
            "<<AGUA_POCO_ARTESIANO_EST>>",
            "<<AGUA_CACIMBA_CISTERNA_POCO>>",
            "<<AGUA_CACIMBA_CISTERNA_POCO_MUN>>",
            "<<AGUA_CACIMBA_CISTERNA_POCO_EST>>",
            "<<AGUA_FONTE_RIO_IGARAPE_RIACHO_CORREGO>>",
            "<<AGUA_FONTE_RIO_IGARAPE_RIACHO_CORREGO_MUN>>",
            "<<AGUA_FONTE_RIO_IGARAPE_RIACHO_CORREGO_EST>>",
            "<<AGUA_NAO_HA>>",
            "<<AGUA_NAO_HA_MUN>>",
            "<<AGUA_NAO_HA_EST>>",
        }:
            if "abs_agua" not in muni_cache:
                muni_cache["abs_agua"] = _calc_abastecimento_agua(municipio, df)
            
            s = muni_cache["abs_agua"]
            def n(k): return str(len(s[k]))
            
            MAP = {
                "<<AGUA_REDE_PUBLICA>>": "rede_publica",
                "<<AGUA_REDE_PUBLICA_MUN>>": "rede_publica_mun",
                "<<AGUA_REDE_PUBLICA_EST>>": "rede_publica_est",
                "<<AGUA_POCO_ARTESIANO>>": "poco_artesiano",
                "<<AGUA_POCO_ARTESIANO_MUN>>": "poco_artesiano_mun",
                "<<AGUA_POCO_ARTESIANO_EST>>": "poco_artesiano_est",
                "<<AGUA_CACIMBA_CISTERNA_POCO>>": "cacimba_etc",
                "<<AGUA_CACIMBA_CISTERNA_POCO_MUN>>": "cacimba_etc_mun",
                "<<AGUA_CACIMBA_CISTERNA_POCO_EST>>": "cacimba_etc_est",
                "<<AGUA_FONTE_RIO_IGARAPE_RIACHO_CORREGO>>": "fonte_etc",
                "<<AGUA_FONTE_RIO_IGARAPE_RIACHO_CORREGO_MUN>>": "fonte_etc_mun",
                "<<AGUA_FONTE_RIO_IGARAPE_RIACHO_CORREGO_EST>>": "fonte_etc_est",
                "<<AGUA_NAO_HA>>": "sem_agua",
                "<<AGUA_NAO_HA_MUN>>": "sem_agua_mun",
                "<<AGUA_NAO_HA_EST>>": "sem_agua_est",
            }
            
            key = MAP.get(tag)
            resultados[tag] = n(key) if key else ""
            continue

         # --------- SISTEMA DE ESGOTAMENTO ----------
        if tag in {
            "<<ESGOTO_REDE_SANITARIA>>",
            "<<ESGOTO_REDE_SANITARIA_MUN>>",
            "<<ESGOTO_REDE_SANITARIA_EST>>",
            "<<ESGOTO_FOSSA_SUMIDOURO>>",
            "<<ESGOTO_FOSSA_SUMIDOURO_MUN>>",
            "<<ESGOTO_FOSSA_SUMIDOURO_EST>>",
            "<<ESGOTO_DESPEJO_INADEQUADO>>",
            "<<ESGOTO_DESPEJO_INADEQUADO_MUN>>",
            "<<ESGOTO_DESPEJO_INADEQUADO_EST>>",
        }:
            if "sist_esg" not in muni_cache:
                muni_cache["sist_esg"] = _calc_sistema_esgotamento(municipio, df)
            
            s = muni_cache["sist_esg"]
            def n(k): return str(len(s[k]))
            
            MAP = {
                "<<ESGOTO_REDE_SANITARIA>>":"sist_conectado",
                "<<ESGOTO_REDE_SANITARIA_MUN>>": "sist_conectado_mun",
                "<<ESGOTO_REDE_SANITARIA_EST>>": "sist_conectado_est",
                "<<ESGOTO_FOSSA_SUMIDOURO>>": "fossa_e_outros",
                "<<ESGOTO_FOSSA_SUMIDOURO_MUN>>": "fossa_e_outros_mun",
                "<<ESGOTO_FOSSA_SUMIDOURO_EST>>": "fossa_e_outros_est",
                "<<ESGOTO_DESPEJO_INADEQUADO>>":"despejo_inadequado",
                "<<ESGOTO_DESPEJO_INADEQUADO_MUN>>":"despejo_inadequado_mun",
                "<<ESGOTO_DESPEJO_INADEQUADO_EST>>":"despejo_inadequado_est",
            }
            
            key = MAP.get(tag)
            resultados[tag] = n(key) if key else ""
            continue

         # --------- BRINQUEDOS PARQUINHO ----------
        if tag in {
            "<<EDUC_INF_PARQUINHO_COND_USO>>",
            "<<EDUC_INF_PARQUINHO_COND_USO_MUN>>",
            "<<EDUC_INF_PARQUINHO_COND_USO_EST>>",
            "<<EDUC_INF_PARQUINHO_SEM_COND>>",
            "<<EDUC_INF_PARQUINHO_SEM_COND_MUN>>",
            "<<EDUC_INF_PARQUINHO_SEM_COND_EST>>",
            "<<EDUC_INF_PARQUINHO_NAO>>",
            "<<EDUC_INF_PARQUINHO_NAO_MUN>>",
            "<<EDUC_INF_PARQUINHO_NAO_EST>>",
        }:
            if "parq_inf" not in muni_cache:
                muni_cache["parq_inf"] = _calc_brinquedos_parquinho(municipio, df)
            
            s = muni_cache["parq_inf"]
            def n(k): return str(len(s[k]))
            
            MAP = {
                "<<EDUC_INF_PARQUINHO_COND_USO>>":"cond",
                "<<EDUC_INF_PARQUINHO_COND_USO_MUN>>":"cond_mun",
                "<<EDUC_INF_PARQUINHO_COND_USO_EST>>":"cond_est",
                "<<EDUC_INF_PARQUINHO_SEM_COND>>":"sem_cond",
                "<<EDUC_INF_PARQUINHO_SEM_COND_MUN>>":"sem_cond_mun",
                "<<EDUC_INF_PARQUINHO_SEM_COND_EST>>":"sem_cond_est",
                "<<EDUC_INF_PARQUINHO_NAO>>":"nao",
                "<<EDUC_INF_PARQUINHO_NAO_MUN>>":"nao_mun",
                "<<EDUC_INF_PARQUINHO_NAO_EST>>":"nao_est",
            }
            
            key = MAP.get(tag)
            resultados[tag] = n(key) if key else ""
            continue

        # --------- PÁTIO ----------
        if tag in {
            "<<EDUC_INF_PATIO_COBERTO_EXCLUSIVO>>",
            "<<EDUC_INF_PATIO_COBERTO_EXCLUSIVO_MUN>>",
            "<<EDUC_INF_PATIO_COBERTO_EXCLUSIVO_EST>>",
            "<<EDUC_INF_PATIO_DESCOBERTO_EXCLUSIVO>>",
            "<<EDUC_INF_PATIO_DESCOBERTO_EXCLUSIVO_MUN>>",
            "<<EDUC_INF_PATIO_DESCOBERTO_EXCLUSIVO_EST>>",
            "<<EDUC_INF_PATIO_COBERTO_COMPART>>",
            "<<EDUC_INF_PATIO_COBERTO_COMPART_MUN>>",
            "<<EDUC_INF_PATIO_COBERTO_COMPART_EST>>",
            "<<EDUC_INF_PATIO_DESCOBERTO_COMPART>>",
            "<<EDUC_INF_PATIO_DESCOBERTO_COMPART_MUN>>",
            "<<EDUC_INF_PATIO_DESCOBERTO_COMPART_EST>>",
            "<<EDUC_INF_PATIO_NAO>>",
            "<<EDUC_INF_PATIO_NAO_MUN>>",
            "<<EDUC_INF_PATIO_NAO_EST>>",
        }:
            if "patio" not in muni_cache:
                muni_cache["patio"] = _calc_infra_patio(municipio, df)
            
            s = muni_cache["patio"]
            def n(k): return str(len(s[k]))
            
            MAP = {
                "<<EDUC_INF_PATIO_COBERTO_EXCLUSIVO>>":"cob_ex",
                "<<EDUC_INF_PATIO_COBERTO_EXCLUSIVO_MUN>>":"cob_ex_mun",
                "<<EDUC_INF_PATIO_COBERTO_EXCLUSIVO_EST>>":"cob_ex_est",
                "<<EDUC_INF_PATIO_DESCOBERTO_EXCLUSIVO>>":"descob_ex",
                "<<EDUC_INF_PATIO_DESCOBERTO_EXCLUSIVO_MUN>>":"descob_ex_mun",
                "<<EDUC_INF_PATIO_DESCOBERTO_EXCLUSIVO_EST>>":"descob_ex_est",
                "<<EDUC_INF_PATIO_COBERTO_COMPART>>":"cob_comp",
                "<<EDUC_INF_PATIO_COBERTO_COMPART_MUN>>":"cob_comp_mun",
                "<<EDUC_INF_PATIO_COBERTO_COMPART_EST>>":"cob_comp_est",
                "<<EDUC_INF_PATIO_DESCOBERTO_COMPART>>":"descob_comp",
                "<<EDUC_INF_PATIO_DESCOBERTO_COMPART_MUN>>":"descob_comp_mun",
                "<<EDUC_INF_PATIO_DESCOBERTO_COMPART_EST>>":"descob_comp_est",
                "<<EDUC_INF_PATIO_NAO>>":"nao",
                "<<EDUC_INF_PATIO_NAO_MUN>>":"nao_mun",
                "<<EDUC_INF_PATIO_NAO_EST>>":"nao_est",
            }
            
            key = MAP.get(tag)
            resultados[tag] = n(key) if key else ""
            continue

        # --------- GESTÃO ALIMENTAÇÃO ----------
        if tag in {
            "<<GESTAO_ALIM_CENTRALIZADA>>",
            "<<GESTAO_ALIM_CENTRALIZADA_MUN>>",
            "<<GESTAO_ALIM_CENTRALIZADA_EST>>",
            "<<GESTAO_ALIM_DESCENTRALIZADA>>",
            "<<GESTAO_ALIM_DESCENTRALIZADA_MUN>>",
            "<<GESTAO_ALIM_DESCENTRALIZADA_EST>>",
            "<<GESTAO_ALIM_SEMIDESCENTRALIZADA>>",
            "<<GESTAO_ALIM_SEMIDESCENTRALIZADA_MUN>>",
            "<<GESTAO_ALIM_SEMIDESCENTRALIZADA_EST>>",
            "<<GESTAO_ALIM_TERCEIRIZADA>>",
            "<<GESTAO_ALIM_TERCEIRIZADA_MUN>>",
            "<<GESTAO_ALIM_TERCEIRIZADA_EST>>",
        }:
            if "gest_alim" not in muni_cache:
                muni_cache["gest_alim"] = _calc_gestao_alim(municipio, df)
            
            s = muni_cache["gest_alim"]
            def n(k): return str(len(s[k]))
            
            MAP = {
                "<<GESTAO_ALIM_CENTRALIZADA>>":"centralizada",
                "<<GESTAO_ALIM_CENTRALIZADA_MUN>>":"centralizada_mun",
                "<<GESTAO_ALIM_CENTRALIZADA_EST>>":"centralizada_est",
                "<<GESTAO_ALIM_DESCENTRALIZADA>>":"descentralizada",
                "<<GESTAO_ALIM_DESCENTRALIZADA_MUN>>":"descentralizada_mun",
                "<<GESTAO_ALIM_DESCENTRALIZADA_EST>>":"descentralizada_est",
                "<<GESTAO_ALIM_SEMIDESCENTRALIZADA>>":"semidescentralizada",
                "<<GESTAO_ALIM_SEMIDESCENTRALIZADA_MUN>>":"semidescentralizada_mun",
                "<<GESTAO_ALIM_SEMIDESCENTRALIZADA_EST>>":"semidescentralizada_est",
                "<<GESTAO_ALIM_TERCEIRIZADA>>":"terceirizada",
                "<<GESTAO_ALIM_TERCEIRIZADA_MUN>>":"terceirizada_mun",
                "<<GESTAO_ALIM_TERCEIRIZADA_EST>>":"terceirizada_est",
            }
            
            key = MAP.get(tag)
            resultados[tag] = n(key) if key else ""
            continue

        # --------- (ALIM) TIPO DE TERCEIRIZAÇÃO ----------
        if tag in {
            "<<GESTAO_ALIM_TERCEIRIZADA_PREP_ESCOLA_EQP_ESCOLA>>",
            "<<GESTAO_ALIM_TERCEIRIZADA_PREP_ESCOLA_EQP_ESCOLA_MUN>>",
            "<<GESTAO_ALIM_TERCEIRIZADA_PREP_ESCOLA_EQP_ESCOLA_EST>>",
            "<<GESTAO_ALIM_TERCEIRIZADA_PREP_ESCOLA_EQP_EMPRESA>>",
            "<<GESTAO_ALIM_TERCEIRIZADA_PREP_ESCOLA_EQP_EMPRESA_MUN>>",
            "<<GESTAO_ALIM_TERCEIRIZADA_PREP_ESCOLA_EQP_EMPRESA_EST>>",
            "<<GESTAO_ALIM_TERCEIRIZADA_ENTREGA_HOTBOX>>",
            "<<GESTAO_ALIM_TERCEIRIZADA_ENTREGA_HOTBOX_MUN>>",
            "<<GESTAO_ALIM_TERCEIRIZADA_ENTREGA_HOTBOX_EST>>",
        }:
            if "alim_terc" not in muni_cache:
                muni_cache["alim_terc"] = _calc_terceirizacao_alim(municipio, df)
            
            s = muni_cache["alim_terc"]
            def n(k): return str(len(s[k]))
            
            MAP = {
                "<<GESTAO_ALIM_TERCEIRIZADA_PREP_ESCOLA_EQP_ESCOLA>>":"eqp_escola",
                "<<GESTAO_ALIM_TERCEIRIZADA_PREP_ESCOLA_EQP_ESCOLA_MUN>>":"eqp_escola_mun",
                "<<GESTAO_ALIM_TERCEIRIZADA_PREP_ESCOLA_EQP_ESCOLA_EST>>":"eqp_escola_est",
                "<<GESTAO_ALIM_TERCEIRIZADA_PREP_ESCOLA_EQP_EMPRESA>>":"eqp_empresa",
                "<<GESTAO_ALIM_TERCEIRIZADA_PREP_ESCOLA_EQP_EMPRESA_MUN>>":"eqp_empresa_mun",
                "<<GESTAO_ALIM_TERCEIRIZADA_PREP_ESCOLA_EQP_EMPRESA_EST>>":"eqp_empresa_est",
                "<<GESTAO_ALIM_TERCEIRIZADA_ENTREGA_HOTBOX>>":"hotbox",
                "<<GESTAO_ALIM_TERCEIRIZADA_ENTREGA_HOTBOX_MUN>>":"hotbox_mun",
                "<<GESTAO_ALIM_TERCEIRIZADA_ENTREGA_HOTBOX_EST>>":"hotbox_est",
            }
            
            key = MAP.get(tag)
            resultados[tag] = n(key) if key else ""
            continue

        # --------- LOCAL ARMAZENAMENTO ----------
        if tag in {
            "<<ARMZ_LOCAL_DESPENSA>>",
            "<<ARMZ_LOCAL_DESPENSA_MUN>>",
            "<<ARMZ_LOCAL_DESPENSA_EST>>",
            "<<ARMZ_LOCAL_ARMARIO_DENTRO_COZINHA>>",
            "<<ARMZ_LOCAL_ARMARIO_DENTRO_COZINHA_MUN>>",
            "<<ARMZ_LOCAL_ARMARIO_DENTRO_COZINHA_EST>>",
            "<<ARMZ_LOCAL_ARMARIO_FORA_COZINHA>>",
            "<<ARMZ_LOCAL_ARMARIO_FORA_COZINHA_MUN>>",
            "<<ARMZ_LOCAL_ARMARIO_FORA_COZINHA_EST>>",
            "<<ARMZ_LOCAL_NAO_HA>>",
            "<<ARMZ_LOCAL_NAO_HA_MUN>>",
            "<<ARMZ_LOCAL_NAO_HA_EST>>",
        }:
            if "local_alim" not in muni_cache:
                muni_cache["local_alim"] = _calc_armazenamento_alim(municipio, df)
            
            s = muni_cache["local_alim"]
            def n(k): return str(len(s[k]))
            
            MAP = {
                "<<ARMZ_LOCAL_DESPENSA>>":"despensa",
                "<<ARMZ_LOCAL_DESPENSA_MUN>>":"despensa_mun",
                "<<ARMZ_LOCAL_DESPENSA_EST>>":"despensa_est",
                "<<ARMZ_LOCAL_ARMARIO_DENTRO_COZINHA>>":"arm_dentro",
                "<<ARMZ_LOCAL_ARMARIO_DENTRO_COZINHA_MUN>>":"arm_dentro_mun",
                "<<ARMZ_LOCAL_ARMARIO_DENTRO_COZINHA_EST>>":"arm_dentro_est",
                "<<ARMZ_LOCAL_ARMARIO_FORA_COZINHA>>":"arm_fora",
                "<<ARMZ_LOCAL_ARMARIO_FORA_COZINHA_MUN>>":"arm_fora_mun",
                "<<ARMZ_LOCAL_ARMARIO_FORA_COZINHA_EST>>":"arm_fora_est",
                "<<ARMZ_LOCAL_NAO_HA>>":"nao_ha",
                "<<ARMZ_LOCAL_NAO_HA_MUN>>":"nao_ha_mun",
                "<<ARMZ_LOCAL_NAO_HA_EST>>":"nao_ha_est",
            }
            
            key = MAP.get(tag)
            resultados[tag] = n(key) if key else ""
            continue

        # --------- LIMITE BERÇÁRIO  ----------
        if tag in {
            "<<BERCARIO_LIMITE_CRIANÇAS_IRREG>>",
            "<<BERCARIO_LIMITE_CRIANÇAS_IRREG_MUN>>",
            "<<BERCARIO_LIMITE_CRIANÇAS_IRREG_EST>>",
        }:
            if "lim_bercario" not in muni_cache:
                muni_cache["lim_bercario"] = _calc_limite_bercario(municipio, df)
            
            s = muni_cache["lim_bercario"]
            def n(k): return str(len(s[k]))
            
            MAP = {
                "<<BERCARIO_LIMITE_CRIANÇAS_IRREG>>":"nao",
                "<<BERCARIO_LIMITE_CRIANÇAS_IRREG_MUN>>":"nao_mun",
                "<<BERCARIO_LIMITE_CRIANÇAS_IRREG_EST>>":"nao_est",
            }
            
            key = MAP.get(tag)
            resultados[tag] = n(key) if key else ""
            continue

        # --------- CONEXÃO REDE ELÉTRICA ----------
        if tag in {
            "<<ENERGIA_SIM_FUNC>>",
            "<<ENERGIA_SIM_FUNC_MUN>>",
            "<<ENERGIA_SIM_FUNC_EST>>",
            "<<ENERGIA_FORA_FUNC>>",
            "<<ENERGIA_FORA_FUNC_MUN>>",
            "<<ENERGIA_FORA_FUNC_EST>>",
            "<<ENERGIA_NAO>>",
            "<<ENERGIA_NAO_MUN>>",
            "<<ENERGIA_NAO_EST>>",
        }:
            if "energia" not in muni_cache:
                muni_cache["energia"] = _calc_energia(municipio, df)
            
            s = muni_cache["energia"]
            def n(k): return str(len(s[k]))
            
            MAP = {
                "<<ENERGIA_SIM_FUNC>>":"em_funcionamento",
                "<<ENERGIA_SIM_FUNC_MUN>>":"em_funcionamento_mun",
                "<<ENERGIA_SIM_FUNC_EST>>":"em_funcionamento_est",
                "<<ENERGIA_FORA_FUNC>>":"fora_funcionamento",
                "<<ENERGIA_FORA_FUNC_MUN>>":"fora_funcionamento_mun",
                "<<ENERGIA_FORA_FUNC_EST>>":"fora_funcionamento_est",
                "<<ENERGIA_NAO>>":"nao",
                "<<ENERGIA_NAO_MUN>>":"nao_mun",
                "<<ENERGIA_NAO_EST>>":"nao_est",
            }
            
            key = MAP.get(tag)
            resultados[tag] = n(key) if key else ""
            continue

        # --------- LOCALIDADE FRALDÁRIOS ----------
        if tag in {
            "<<FRALDARIO_LOCAL_DENTRO_BERCARIO>>",
            "<<FRALDARIO_LOCAL_DENTRO_BERCARIO_MUN>>",
            "<<FRALDARIO_LOCAL_DENTRO_BERCARIO_EST>>",
            "<<FRALDARIO_LOCAL_SEPARADO>>",
            "<<FRALDARIO_LOCAL_SEPARADO_MUN>>",
            "<<FRALDARIO_LOCAL_SEPARADO_EST>>",
        }:
            if "local_fral" not in muni_cache:
                muni_cache["local_fral"] = _calc_localidade_fraldario(municipio, df)
            
            s = muni_cache["local_fral"]
            def n(k): return str(len(s[k]))
            
            MAP = {
                "<<FRALDARIO_LOCAL_DENTRO_BERCARIO>>": "local_dentro",
                "<<FRALDARIO_LOCAL_DENTRO_BERCARIO_MUN>>": "local_dentro_mun",
                "<<FRALDARIO_LOCAL_DENTRO_BERCARIO_EST>>": "local_dentro_est",
                "<<FRALDARIO_LOCAL_SEPARADO>>": "local_separado",
                "<<FRALDARIO_LOCAL_SEPARADO_MUN>>": "local_separado_mun",
                "<<FRALDARIO_LOCAL_SEPARADO_EST>>": "local_separado_est",
            }
            
            key = MAP.get(tag)
            resultados[tag] = n(key) if key else ""
            continue

        # --------- CARDÁPIO ESPECIAL  ----------
        if tag in {
            "<<CARDAPIO_NEEDS_ESPECIAL_COM_CARDAPIO>>",
            "<<CARDAPIO_NEEDS_ESPECIAL_COM_CARDAPIO_MUN>>",
            "<<CARDAPIO_NEEDS_ESPECIAL_COM_CARDAPIO_EST>>",
            "<<CARDAPIO_NEEDS_ESPECIAL_SEM_CARDAPIO>>",
            "<<CARDAPIO_NEEDS_ESPECIAL_SEM_CARDAPIO_MUN>>",
            "<<CARDAPIO_NEEDS_ESPECIAL_SEM_CARDAPIO_EST>>",
            "<<CARDAPIO_NEEDS_NAO>>",
            "<<CARDAPIO_NEEDS_NAO_MUN>>",
            "<<CARDAPIO_NEEDS_NAO_EST>>",
        }:
            if "card_especial" not in muni_cache:
                muni_cache["card_especial"] = _calc_cardapio_especial(municipio, df)
            
            s = muni_cache["card_especial"]
            def n(k): return str(len(s[k]))
            
            MAP = {
                "<<CARDAPIO_NEEDS_ESPECIAL_COM_CARDAPIO>>":"sim",
                "<<CARDAPIO_NEEDS_ESPECIAL_COM_CARDAPIO_MUN>>":"sim_mun",
                "<<CARDAPIO_NEEDS_ESPECIAL_COM_CARDAPIO_EST>>":"sim_est",
                "<<CARDAPIO_NEEDS_ESPECIAL_SEM_CARDAPIO>>":"sim_irr",
                "<<CARDAPIO_NEEDS_ESPECIAL_SEM_CARDAPIO_MUN>>":"sim_irr_mun",
                "<<CARDAPIO_NEEDS_ESPECIAL_SEM_CARDAPIO_EST>>":"sim_irr_est",
                "<<CARDAPIO_NEEDS_NAO>>":"nao",
                "<<CARDAPIO_NEEDS_NAO_MUN>>":"nao_mun",
                "<<CARDAPIO_NEEDS_NAO_EST>>":"nao_est",
            }
            
            key = MAP.get(tag)
            resultados[tag] = n(key) if key else ""
            continue

        # --------- REFEIÇÃO PREPARADA  ----------
        if tag in {
            "<<CARDAPIO_VISITA_REFEICAO_SERVIDA>>",
            "<<CARDAPIO_VISITA_REFEICAO_SERVIDA_MUN>>",
            "<<CARDAPIO_VISITA_REFEICAO_SERVIDA_EST>>",
            "<<CARDAPIO_VISITA_REFEICAO_PREPARADA>>",
            "<<CARDAPIO_VISITA_REFEICAO_PREPARADA_MUN>>",
            "<<CARDAPIO_VISITA_REFEICAO_PREPARADA_EST>>",
            "<<CARDAPIO_VISITA_REFEICAO_NAO>>",
            "<<CARDAPIO_VISITA_REFEICAO_NAO_MUN>>",
            "<<CARDAPIO_VISITA_REFEICAO_NAO_EST>>",
        }:
            if "visita_refeicao" not in muni_cache:
                muni_cache["visita_refeicao"] = _calc_refeicao_servida(municipio, df)
            
            s = muni_cache["visita_refeicao"]
            def n(k): return str(len(s[k]))
            
            MAP = {
                "<<CARDAPIO_VISITA_REFEICAO_SERVIDA>>":"serv",
                "<<CARDAPIO_VISITA_REFEICAO_SERVIDA_MUN>>":"serv_mun",
                "<<CARDAPIO_VISITA_REFEICAO_SERVIDA_EST>>":"serv_est",
                "<<CARDAPIO_VISITA_REFEICAO_PREPARADA>>":"prep",
                "<<CARDAPIO_VISITA_REFEICAO_PREPARADA_MUN>>":"prep_mun",
                "<<CARDAPIO_VISITA_REFEICAO_PREPARADA_EST>>":"prep_est",
                "<<CARDAPIO_VISITA_REFEICAO_NAO>>":"nao",
                "<<CARDAPIO_VISITA_REFEICAO_NAO_MUN>>":"nao_mun",
                "<<CARDAPIO_VISITA_REFEICAO_NAO_EST>>":"nao_est",
            }
            
            key = MAP.get(tag)
            resultados[tag] = n(key) if key else ""
            continue

          # --------- CARDÁPIO CONFORME  ----------
        if tag in {
            "<<CARDAPIO_CONFORME_TODOS_ITENS>>",
            "<<CARDAPIO_CONFORME_TODOS_ITENS_MUN>>",
            "<<CARDAPIO_CONFORME_TODOS_ITENS_EST>>",
            "<<CARDAPIO_CONFORME_APENAS_ALGUNS>>",
            "<<CARDAPIO_CONFORME_APENAS_ALGUNS_MUN>>",
            "<<CARDAPIO_CONFORME_APENAS_ALGUNS_EST>>",
            "<<CARDAPIO_CONFORME_NENHUM>>",
            "<<CARDAPIO_CONFORME_NENHUM_MUN>>",
            "<<CARDAPIO_CONFORME_NENHUM_EST>>",
        }:
            if "cardapio_conforme" not in muni_cache:
                muni_cache["cardapio_conforme"] = _calc_cardapio_conforme(municipio, df)
            
            s = muni_cache["cardapio_conforme"]
            def n(k): return str(len(s[k]))
            
            MAP = {
                "<<CARDAPIO_CONFORME_TODOS_ITENS>>":"conf_todos",
                "<<CARDAPIO_CONFORME_TODOS_ITENS_MUN>>":"conf_todos_mun",
                "<<CARDAPIO_CONFORME_TODOS_ITENS_EST>>":"conf_todos_est",
                "<<CARDAPIO_CONFORME_APENAS_ALGUNS>>":"conf_alguns",
                "<<CARDAPIO_CONFORME_APENAS_ALGUNS_MUN>>":"conf_alguns_mun",
                "<<CARDAPIO_CONFORME_APENAS_ALGUNS_EST>>":"conf_alguns_est",
                "<<CARDAPIO_CONFORME_NENHUM>>":"nenhum",
                "<<CARDAPIO_CONFORME_NENHUM_MUN>>":"nenhum_mun",
                "<<CARDAPIO_CONFORME_NENHUM_EST>>":"nenhum_est",
            }
            
            key = MAP.get(tag)
            resultados[tag] = n(key) if key else ""
            continue

        # --------- SALAS MULTISSERIADAS ----------
        if tag in {
            "<<SALAS_MULTISSERIADAS>>",
            "<<SALAS_MULTISSERIADAS_MUN>>",
            "<<SALAS_MULTISSERIADAS_EST>>",
        }:
            if "sala_multisseriada" not in muni_cache:
                muni_cache["sala_multisseriada"] = _calc_salas_multisseriadas(municipio, df)
            
            s = muni_cache["sala_multisseriada"]
            def n(k): return str(len(s[k]))
            
            MAP = {
                "<<SALAS_MULTISSERIADAS>>":"sim",
                "<<SALAS_MULTISSERIADAS_MUN>>":"sim_mun",
                "<<SALAS_MULTISSERIADAS_EST>>":"sim_est",
            }
            
            key = MAP.get(tag)
            resultados[tag] = n(key) if key else ""
            continue

        if tag.startswith("<<RESERVATORIO"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            # Muda o sufixo para bater com a questão
            sufixo = tag.replace("<<", "").replace(">>", "").replace("RESERVATORIO","")
            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"reservatorio_agua_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_reservatorio_agua_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue


        if tag.startswith("<<ARMZ_LOCAL_IRREG"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            # Muda o sufixo para bater com a questão
            sufixo = tag.replace("<<", "").replace(">>", "").replace("ARMZ_LOCAL_IRREG", "").replace("ENFERRUJADA","ENFERRUJADO").replace("INAPROPRIADA","NAO_APROPRIADAS").replace("DIRETO","DIRETAMENTE").replace("SEM","NAO_POSSUI")

            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"irr_armazenamento_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_irr_armazenamento_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

        if tag.startswith("<<COZ_OUTROS"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("COZ_OUTROS", "")

            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"cozinha_outros_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_cozinha_outros_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

        if tag.startswith("<<BIB_EXC_INFRA"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("BIB_EXC_INFRA", "")

            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"irr_bib_exc_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_irr_biblioteca_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

        if tag.startswith("<<SL_EXC_INFRA"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("SL_EXC_INFRA", "")

            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"irr_sl_exc_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_irr_sala_leitura_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

        if tag.startswith("<<BIBLI_SL_COMP_INFRA"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("BIBLI_SL_COMP_INFRA", "")

            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"irr_bibsl_exc_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_irr_bib_compart_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

        if tag.startswith("<<EI_SAN_EXC_AGUA"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("EI_SAN_EXC_AGUA", "")

            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"ei_san_exc_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_agua_san_infantil_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

        if tag.startswith("<<EI_SAN_EXC_INFRA"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("EI_SAN_EXC_INFRA", "")

            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"irr_ei_san_exc_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_irr_san_infantil_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue


        if tag.startswith("<<ARMZ_UP"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("ARMZ_UP", "")

            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"ultraprocessados_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_ultraprocessados_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

        if tag.startswith("<<UP_CONSUMO"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("UP_CONSUMO", "")

            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"cons_ultraprocessados_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_cons_ultraprocessados_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

        if tag.startswith("<<UP_DIRETRIZ"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("UP_DIRETRIZ", "")

            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"dir_ultraprocessados_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_dir_ultraprocessados_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

        if tag.startswith("<<UP_COMERCIO"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("UP_COMERCIO", "")

            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"venda_ultraprocessados_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_venda_ultraprocessados_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue


        if tag.startswith("<<ARMZ_CONGELADOS"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("ARMZ_CONGELADOS", "")

            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"armz_congelados_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_armz_congelados_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

        if tag.startswith("<<ARMZ_CONG_IRREG"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("ARMZ_CONGELADOS_IRREG", "")

            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"irr_congelados_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_irr_congelados_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue
        tags_ass_cardapio = [
            "<<CARDAPIO_ASS_RT_E_FIXADO>>",
            "<<CARDAPIO_ASS_RT_E_FIXADO_MUN>>",
            "<<CARDAPIO_ASS_RT_E_FIXADO_EST>>",
            "<<CARDAPIO_NAO_ASSINADO_MAS_FIXADO>>",
            "<<CARDAPIO_NAO_ASSINADO_MAS_FIXADO_MUN>>",
            "<<CARDAPIO_NAO_ASSINADO_MAS_FIXADO_EST>>",
            "<<CARDAPIO_ASSINADO_NAO_FIXADO>>",
            "<<CARDAPIO_ASSINADO_NAO_FIXADO_MUN>>",
            "<<CARDAPIO_ASSINADO_NAO_FIXADO_EST>>",
            "<<CARDAPIO_NAO_ASS_NAO_FIXADO>>",
            "<<CARDAPIO_NAO_ASS_NAO_FIXADO_MUN>>",
            "<<CARDAPIO_NAO_ASS_NAO_FIXADO_EST>>",
            "<<CARDAPIO_NAO_EXISTE>>",
            "<<CARDAPIO_NAO_EXISTE_MUN>>",
            "<<CARDAPIO_NAO_EXISTE_EST>>",
        ]

        if tag.startswith("<<CARDAPIO") and tag in tags_ass_cardapio:
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("CARDAPIO", "").replace("NAO_EXISTE","INEXISTENTE").replace("ASS_","ASSINADO_").replace("_NAO_ASSINADO_NAO_FIXADO","NAONAO").replace("_NAO_ASSINADO_MAS_FIXADO","NAOMAS")
            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"cardapio_assinado_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_cardapio_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue


        if tag.startswith("<<COZ_INFRA"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("COZ_INFRA", "")

            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"patio_educ_inf_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_infra_cozinha_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue


        if tag.startswith("<<NUM_ESCOLAS"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("NUM_ESCOLAS", "").replace("INADEQ", "")
            
            # Determina qual esfera contar
            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"acesso_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_acessibilidade_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

        

        if tag.startswith("<<SALAS_IRREGULARES"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("SALAS_IRREGULARES", "")
            
            # Determina qual esfera contar
            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"sala_irr_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_irregularidades_salas_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

        if tag.startswith("<<LACTARIO_ITEM"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("LACTARIO_ITEM", "")
            
            # Determina qual esfera contar
            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"item_lactario_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_item_lactario_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

        if tag.startswith("<<LACTARIO_INFRA"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("LACTARIO_INFRA", "")
            
            # Determina qual esfera contar
            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"irr_lactario_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_irregularidades_lactario_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

        if tag.startswith("<<FRALDARIO_ITEM"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("FRALDARIO_ITEM", "")
            
            # Determina qual esfera contar
            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"item_fraldario_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_item_fraldario_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue
        
        if tag.startswith("<<FRALDARIO_INFRA"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            # Substitui abreviações comuns
            sufixo = tag.replace("<<", "").replace(">>", "").replace("FRALDARIO_INFRA", "").replace("VENT", "VENTILAÇÃO").replace("ILUM","ILUMINAÇÃO")
            
            # Determina qual esfera contar
            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"irr_fraldario_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_irregularidades_fraldario_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

        deps_infantil = [
            "<<DEP_SANITARIO_ED_INFANTIL>>",
            "<<DEP_SANITARIO_ED_INFANTIL_MUN>>",
            "<<DEP_SANITARIO_ED_INFANTIL_EST>>",
            "<<DEP_BERCARIO>>",
            "<<DEP_BERCARIO_MUN>>",
            "<<DEP_BERCARIO_EST>>",
            "<<DEP_FRALDARIO>>",
            "<<DEP_FRALDARIO_MUN>>",
            "<<DEP_FRALDARIO_EST>>",
            "<<DEP_LACTARIO>>",
            "<<DEP_LACTARIO_MUN>>",
            "<<DEP_LACTARIO_EST>>",
            "<<DEP_PARQUE_INFANTIL>>",
            "<<DEP_PARQUE_INFANTIL_MUN>>",
            "<<DEP_PARQUE_INFANTIL_EST>>",
            "<<DEP_LAVANDERIA_AREA_LAVAGEM>>",
            "<<DEP_LAVANDERIA_AREA_LAVAGEM_MUN>>",
            "<<DEP_LAVANDERIA_AREA_LAVAGEM_EST>>",
        ]

        if tag.startswith("<<DEP_") and tag not in deps_infantil:
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("DEP_", "")
            
            
            # Determina qual esfera contar
            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"dep_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_dependencias_por_tag(municipio, df, tag_base)
                
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

        if tag.startswith("<<DEP_") and tag in deps_infantil:
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("DEP_", "")
            
            
            # Determina qual esfera contar
            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"dep_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_deps_inf_por_tag(municipio, df, tag_base)
                
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

        if tag.startswith("<<LIXO"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("LIXO", "")
            
            # Determina qual esfera contar
            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"dest_lixo_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_destinacao_lixo_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

        # Modificado para acomodar as tags ventilação natural e ventilação mecânica
        if tag.startswith("<<BERCARIO_INFRA"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("BERCARIO_INFRA", "").replace("VENT_NAT","VENTILACAO_NATURAL").replace("VENT_MEC","VENTILADORES")
            
            # Determina qual esfera contar
            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"irr_bercario_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_irregularidades_bercario_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue

         # Modificado para acomodar as tags ventilação natural e ventilação mecânica
        if tag.startswith("<<REFEITORIO_INFRA"):
            # Remove << >> e o prefixo DEPENDENCIAS_
            sufixo = tag.replace("<<", "").replace(">>", "").replace("REFEITORIO_INFRA", "")
            
            # Determina qual esfera contar
            if sufixo.endswith("_MUN"):
                esfera_alvo = "mun"
                tag_base = sufixo[:-4]  # Remove "_MUN"
            elif sufixo.endswith("_EST"):
                esfera_alvo = "est"
                tag_base = sufixo[:-4]  # Remove "_EST"
            else:
                esfera_alvo = "total"
                tag_base = sufixo
            
            # Usa cache para evitar recalcular a mesma dependência
            cache_key = f"irr_refeitorio_{tag_base}"
            if cache_key not in muni_cache:
                # Calcula pela primeira vez
                muni_cache[cache_key] = _calc_irr_refeitorio_por_tag(municipio, df, tag_base)
            
            # Retorna a contagem da esfera solicitada
            resultados[tag] = str(len(muni_cache[cache_key][esfera_alvo]))
            continue
        
        # Tag não reconhecida
        resultados[tag] = ""
    
    return resultados

# ========================= ESCRITA DA PLANILHA =========================
def _escrever_planilha(saida: Path, tags: List[str], municipios: List[str], 
                       municipio_alvo: str, resultados: Dict[str, str]):
    """Escreve a planilha de saída com os resultados."""
    print(f"Gerando planilha: {saida}")
    
    wb = Workbook()
    ws_calc = wb.active
    ws_calc.title = "ListaTagsCalculadas"
    
    # Cabeçalhos
    ws_calc["A1"] = "Município:"
    ws_calc["B1"] = "=ListaTagsSimples!B1"

    ws_calc["A2"] = "Tags a serem substituídas"
    ws_calc["B2"] = ""
    ws_calc["C2"] = "Script"
    ws_calc["D2"] = "Descrição"

    
    # Cabeçalho de municípios (E2 →)
    start_col = 5
    for j, m in enumerate(municipios, start=start_col):
        ws_calc.cell(row=2, column=j, value=m)

    # Tags (coluna A a partir de A3)
    for i, tag in enumerate(tags, start=3):
        ws_calc.cell(row=i, column=1, value=tag)

    # Fórmula B3 (igual à sua)
    if tags:
        ws_calc["B3"] = (
            "=LET("
            "mun, ListaTagsSimples!B1,"
            "headers, E2:2,"
            "matriz, E3:AAA,"
            "col, CORRESP(mun, headers, 0),"
            "ÍNDICE(matriz, , col)"
            ")"
        )

    # Preenche a coluna do município-alvo
    try:
        col_dest = start_col + municipios.index(municipio_alvo)
    except ValueError:
        municipios.append(municipio_alvo)
        col_dest = start_col + len(municipios) - 1
        ws_calc.cell(row=2, column=col_dest, value=municipio_alvo)

    for i, tag in enumerate(tags, start=3):
        ws_calc.cell(row=i, column=col_dest, value=resultados.get(tag, ""))

    # Larguras
    ws_calc.column_dimensions["A"].width = 48
    ws_calc.column_dimensions["B"].width = 24
    ws_calc.column_dimensions["C"].width = 18
    ws_calc.column_dimensions["D"].width = 60
    for j in range(start_col, start_col + max(1, len(municipios))):
        ws_calc.column_dimensions[get_column_letter(j)].width = 22

    # Aba ListaTagsSimples (B1 define o município)
    ws_simple = wb.create_sheet("ListaTagsSimples")
    ws_simple["A1"] = "Selecione o município (B1):"
    ws_simple["B1"] = municipio_alvo
    ws_simple["A3"] = "Municípios (ordem alfabética):"
    for i, m in enumerate(municipios, start=4):
        ws_simple.cell(row=i, column=1, value=m)

    wb.save(saida)
    print("Planilha salva com sucesso!")

# ========================= MAIN =========================
def main():
    ap = argparse.ArgumentParser(
        description="Gera TagsCalculadas.xlsx com TODOS os cálculos (escolas visitadas + vigilância sanitária)."
    )
    ap.add_argument("--basedados", required=True, type=Path, help="Caminho para BaseDados.xlsx")
    ap.add_argument("--tags", required=True, type=Path, help="Caminho para TAGS_a_calcular.txt")
    ap.add_argument("--municipio", required=True, help="Nome do município alvo")
    ap.add_argument("--saida", required=True, type=Path, help="Caminho para TagsCalculadas.xlsx")
    args = ap.parse_args()
    
    # Carrega dados
    df = _carregar_base_dados(args.basedados)
    tags = _carregar_tags(args.tags)
    municipios = _extrair_municipios(df)
    
    # Calcula tags
    resultados = calcular_tags_para_municipio(args.municipio, df, tags)
    
    # Escreve planilha
    _escrever_planilha(args.saida, tags, municipios, args.municipio, resultados)
    
    print(f"Processamento concluído com sucesso!")
    print(f"  Município: {args.municipio}")
    print(f"  Tags processadas: {len(tags)}")
    print(f"  Arquivo gerado: {args.saida}")

if __name__ == "__main__":
    main()
